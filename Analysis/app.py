import streamlit as st
import pandas as pd

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="NSE F&O Analysis",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS — Dashboard Theme ───────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ── Base ── */
html, body, [data-testid="stAppViewContainer"], .main {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}
.main { background: #f1f5f9 !important; }
.main .block-container {
    padding-top: 0 !important;
    padding-left: 0 !important;
    padding-right: 0 !important;
    padding-bottom: 0 !important;
    max-width: 100% !important;
}

/* ── Remove Streamlit top header bar ── */
[data-testid="stHeader"] { display: none !important; height: 0 !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
header { display: none !important; height: 0 !important; }
:root { --header-height: 0rem !important; }

/* ── Kill all top padding/margin on every container layer ── */
section.main > div { padding-top: 0 !important; margin-top: 0 !important; }
.appview-container .main .block-container { padding-top: 0 !important; }
[data-testid="stMainBlockContainer"] { padding-top: 0 !important; }
div.block-container { padding-top: 0 !important; margin-top: 0 !important; }
.stApp > div:first-child { padding-top: 0 !important; margin-top: 0 !important; }

/* ── Pull navbar flush to top ── */
.top-navbar { margin-top: -2rem !important; }

/* ── Tighten vertical spacing between elements ── */
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stVerticalBlock"] > div { gap: 0 !important; }

/* ── Sidebar ── */
[data-testid="stSidebar"] {
    background: #0f172a !important;
    border-right: 1px solid #1e293b !important;
}
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] .stCaption { color: #94a3b8 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #f1f5f9 !important; }
[data-testid="stSidebar"] input {
    background: #1e293b !important;
    border: 1px solid #334155 !important;
    color: #f1f5f9 !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] details summary,
[data-testid="stSidebar"] details p { color: #94a3b8 !important; }

/* ── Top navbar ── */
.top-navbar {
    background: linear-gradient(135deg, #0f172a 0%, #1e3a5f 100%);
    padding: 1.1rem 2.2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-bottom: 0;
}
.nav-title {
    font-size: 1.45rem;
    font-weight: 700;
    color: #f8fafc;
    letter-spacing: -0.02em;
    margin: 0;
}
.nav-subtitle {
    font-size: 0.78rem;
    color: #94a3b8;
    margin-top: 3px;
}
.nav-badge {
    background: #1d4ed8;
    color: #bfdbfe;
    padding: 5px 14px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.03em;
}

/* ── Content wrapper ── */
.content-wrap { padding: 0.6rem 2rem; }

/* ── Stat cards ── */
div[data-testid="metric-container"] {
    background: white;
    border-radius: 10px;
    padding: 1rem 1.2rem !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    border-left: 4px solid #3b82f6;
}
div[data-testid="metric-container"] label { color: #64748b !important; font-size: 0.72rem !important; text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500 !important; }
div[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #0f172a !important; font-size: 1.4rem !important; font-weight: 700 !important; }

/* ── Section headers ── */
.section-header {
    background: #0f172a;
    padding: 0.85rem 1.4rem;
    border-radius: 10px 10px 0 0;
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-top: 0.5rem;
}
.section-header-title {
    color: #f8fafc;
    font-size: 0.95rem;
    font-weight: 600;
    margin: 0;
    letter-spacing: -0.01em;
}
.section-badge {
    background: #1d4ed8;
    color: #bfdbfe;
    font-size: 0.7rem;
    padding: 3px 10px;
    border-radius: 12px;
    font-weight: 500;
}
.section-body {
    background: white;
    border-radius: 0 0 10px 10px;
    padding: 1.2rem 1.4rem;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    margin-bottom: 1.5rem;
}

/* ── Tables ── */
div[data-testid="stDataFrame"] {
    border-radius: 8px !important;
    overflow: hidden !important;
    border: 1px solid #e2e8f0 !important;
}

/* ── Buttons ── */
.stButton > button {
    background: #1d4ed8 !important;
    color: white !important;
    border: none !important;
    border-radius: 7px !important;
    font-family: 'Inter', sans-serif !important;
    font-weight: 500 !important;
    font-size: 0.85rem !important;
    padding: 0.45rem 1.2rem !important;
    transition: background 0.2s !important;
}
.stButton > button:hover { background: #1e40af !important; }

/* ── Inputs ── */
.stTextInput input {
    border-radius: 7px !important;
    border: 1.5px solid #e2e8f0 !important;
    background: #f8fafc !important;
    font-size: 0.87rem !important;
    font-family: 'Inter', sans-serif !important;
}
.stTextInput input:focus {
    border-color: #3b82f6 !important;
    box-shadow: 0 0 0 3px rgba(59,130,246,0.15) !important;
    background: white !important;
}

/* ── Info / warning boxes ── */
div[data-testid="stInfo"] {
    background: #eff6ff !important;
    border: 1px solid #bfdbfe !important;
    border-radius: 8px !important;
    color: #1e40af !important;
}
div[data-testid="stWarning"] {
    border-radius: 8px !important;
}

/* ── Caption ── */
.stCaption p { color: #64748b !important; font-size: 0.77rem !important; }

/* ── Progress bar ── */
div[data-testid="stProgressBar"] > div { background: #3b82f6 !important; }

/* ── Divider ── */
hr { border-color: #e2e8f0 !important; margin: 1.2rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:1.2rem 0.5rem 1rem; border-bottom:1px solid #1e293b; margin-bottom:1rem;">
        <div style="font-size:1.1rem;font-weight:700;color:#f8fafc;letter-spacing:-0.02em;">📊 F&amp;O Analysis</div>
        <div style="font-size:0.72rem;color:#64748b;margin-top:3px;">NSE Derivatives Screener</div>
    </div>
    """, unsafe_allow_html=True)

# ── Top Navbar ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="top-navbar">
    <div>
        <div class="nav-title">📊 NSE F&amp;O Analysis</div>
        <div class="nav-subtitle">Futures &amp; Options Screener — <span style="font-size:0.72rem;opacity:0.7;">powered by yfinance</span></div>
    </div>
    <span class="nav-badge">F&amp;O Live</span>
</div>
<div class="content-wrap" style="padding-bottom:0">
""", unsafe_allow_html=True)

# ── Company name lookup: NSE symbol → display name ───────────────────────────
_COMPANY_NAMES = {
    "ABB":"ABB India","POWERINDIA":"Hitachi Energy India","ADANIENT":"Adani Enterprises",
    "ADANIGREEN":"Adani Green Energy","ADANIPORTS":"Adani Ports & SEZ","ADANIPOWER":"Adani Power",
    "ADANIENSOL":"Adani Energy Solutions","ABCAPITAL":"Aditya Birla Capital","ALKEM":"Alkem Laboratories",
    "GVT&D":"GE Vernova T&D India","AMBUJACEM":"Ambuja Cements","AMBER":"Amber Enterprises India",
    "ANGELONE":"Angel One","APLAPOLLO":"APL Apollo Tubes","APOLLOHOSP":"Apollo Hospitals",
    "ASHOKLEY":"Ashok Leyland","ASIANPAINT":"Asian Paints","ASTRAL":"Astral",
    "AUROPHARMA":"Aurobindo Pharma","AUBANK":"AU Small Finance Bank","DMART":"Avenue Supermarts (D-Mart)",
    "AXISBANK":"Axis Bank","BAJAJ-AUTO":"Bajaj Auto","BAJAJFINSV":"Bajaj Finserv",
    "BAJFINANCE":"Bajaj Finance","BAJAJHLDNG":"Bajaj Holdings","BANDHANBNK":"Bandhan Bank",
    "BANKBARODA":"Bank of Baroda","BANKINDIA":"Bank of India","BHARTIARTL":"Bharti Airtel",
    "BDL":"Bharat Dynamics","BEL":"Bharat Electronics","BHARATFORG":"Bharat Forge",
    "INDUSTOWER":"Indus Towers","BPCL":"BPCL","BHEL":"BHEL","BIOCON":"Biocon",
    "BLUESTARCO":"Blue Star","BOSCHLTD":"Bosch India","BRITANNIA":"Britannia Industries",
    "BSE":"BSE","ZYDUSLIFE":"Zydus Lifesciences","CANBK":"Canara Bank","CDSL":"CDSL",
    "CHOLAFIN":"Cholamandalam Investment & Finance","CIPLA":"Cipla","COALINDIA":"Coal India",
    "COCHINSHIP":"Cochin Shipyard","COLPAL":"Colgate-Palmolive India","CAMS":"CAMS",
    "CONCOR":"Container Corp of India","CROMPTON":"Crompton Greaves Consumer","CGPOWER":"CG Power",
    "CUMMINSIND":"Cummins India","DABUR":"Dabur India","DELHIVERY":"Delhivery",
    "DIVISLAB":"Divi's Laboratories","DIXON":"Dixon Technologies","DLF":"DLF",
    "DRREDDY":"Dr. Reddy's Laboratories","EICHERMOT":"Eicher Motors","FEDERALBNK":"Federal Bank",
    "FORTIS":"Fortis Healthcare","FORCEMOT":"Force Motors","NYKAA":"Nykaa (FSN E-Commerce)",
    "GAIL":"GAIL India","GLENMARK":"Glenmark Pharmaceuticals","GMRAIRPORT":"GMR Airports",
    "GODREJCP":"Godrej Consumer Products","GODFRYPHLP":"Godfrey Phillips India",
    "GODREJPROP":"Godrej Properties","GRASIM":"Grasim Industries","HAVELLS":"Havells India",
    "HCLTECH":"HCL Technologies","HDFCAMC":"HDFC AMC","HDFCBANK":"HDFC Bank",
    "HDFCLIFE":"HDFC Life Insurance","HEROMOTOCO":"Hero MotoCorp","HAL":"Hindustan Aeronautics (HAL)",
    "HINDALCO":"Hindalco Industries","HINDUNILVR":"Hindustan Unilever","HINDPETRO":"HPCL",
    "HINDZINC":"Hindustan Zinc","HYUNDAI":"Hyundai Motor India","ICICIBANK":"ICICI Bank",
    "ICICIGI":"ICICI Lombard GI","ICICIPRULI":"ICICI Prudential Life","IDEA":"Vodafone Idea",
    "IDFCFIRSTB":"IDFC First Bank","360ONE":"360 ONE WAM","INDUSINDBK":"IndusInd Bank",
    "IEX":"Indian Energy Exchange","INDHOTEL":"Indian Hotels (Taj)","INDIANB":"Indian Bank",
    "IOC":"Indian Oil Corp","IRFC":"IRFC","IREDA":"IREDA","NAUKRI":"Info Edge (Naukri)",
    "INFY":"Infosys","INOXWIND":"INOX Wind","INDIGO":"IndiGo (InterGlobe Aviation)","ITC":"ITC",
    "JINDALSTEL":"Jindal Steel & Power","JIOFIN":"Jio Financial Services","JSWENERGY":"JSW Energy",
    "JSWSTEEL":"JSW Steel","JUBLFOOD":"Jubilant FoodWorks","KALYANKJIL":"Kalyan Jewellers",
    "KAYNES":"Kaynes Technology","KEI":"KEI Industries","KFINTECH":"KFin Technologies",
    "KOTAKBANK":"Kotak Mahindra Bank","KPITTECH":"KPIT Technologies","LT":"Larsen & Toubro",
    "LAURUSLABS":"Laurus Labs","LICI":"LIC India","LICHSGFIN":"LIC Housing Finance",
    "LTF":"L&T Finance","LTM":"LTIMindtree","LUPIN":"Lupin","LODHA":"Lodha Developers",
    "M&M":"Mahindra & Mahindra","MANAPPURAM":"Manappuram Finance","MANKIND":"Mankind Pharma",
    "MARICO":"Marico","MARUTI":"Maruti Suzuki","MFSL":"Max Financial Services",
    "MAXHEALTH":"Max Healthcare","MAZDOCK":"Mazagon Dock Shipbuilders","MCX":"MCX",
    "UNOMINDA":"UNO Minda","MOTILALOFS":"Motilal Oswal Financial Services",
    "MOTHERSON":"Motherson Sumi Systems","MPHASIS":"Mphasis","MUTHOOTFIN":"Muthoot Finance",
    "NATIONALUM":"National Aluminium","NMDC":"NMDC","NBCC":"NBCC India",
    "NESTLEIND":"Nestlé India","NHPC":"NHPC","COFORGE":"Coforge","NTPC":"NTPC",
    "OBEROIRLTY":"Oberoi Realty","DALBHARAT":"Dalmia Bharat","OIL":"Oil India",
    "PAYTM":"Paytm (One 97 Communications)","ONGC":"ONGC","OFSS":"Oracle Financial Services",
    "PAGEIND":"Page Industries","POLICYBZR":"PB Fintech (Policybazaar)",
    "PERSISTENT":"Persistent Systems","PETRONET":"Petronet LNG","PGEL":"PG Electroplast",
    "PHOENIXLTD":"Phoenix Mills","PIDILITIND":"Pidilite Industries","PIIND":"PI Industries",
    "PNBHOUSING":"PNB Housing Finance","POLYCAB":"Polycab India","PFC":"Power Finance Corp (PFC)",
    "POWERGRID":"Power Grid Corp","PREMIERENE":"Premier Energies","PRESTIGE":"Prestige Estates Projects",
    "PNB":"Punjab National Bank","RADICO":"Radico Khaitan","RVNL":"Rail Vikas Nigam (RVNL)",
    "RBLBANK":"RBL Bank","RELIANCE":"Reliance Industries","NAM-INDIA":"Nippon India AMC",
    "PATANJALI":"Patanjali Foods","RECLTD":"REC","SAIL":"Steel Authority of India (SAIL)",
    "SBICARD":"SBI Cards","SBILIFE":"SBI Life Insurance","SHREECEM":"Shree Cement",
    "SHRIRAMFIN":"Shriram Finance","SIEMENS":"Siemens India","SOLARINDS":"Solar Industries India",
    "SONACOMS":"Sona BLW Precision Forgings","SRF":"SRF","SBIN":"State Bank of India",
    "SUNPHARMA":"Sun Pharma","SUPREMEIND":"Supreme Industries","SUZLON":"Suzlon Energy",
    "SWIGGY":"Swiggy","TATAELXSI":"Tata Elxsi","TATACONSUM":"Tata Consumer Products",
    "TATAMOTORS":"Tata Motors","TATAPOWER":"Tata Power","TATASTEEL":"Tata Steel",
    "TCS":"Tata Consultancy Services","TECHM":"Tech Mahindra","TITAN":"Titan Company",
    "TORNTPHARM":"Torrent Pharmaceuticals","TRENT":"Trent","TIINDIA":"Tube Investments of India",
    "TVSMOTOR":"TVS Motor","ULTRACEMCO":"UltraTech Cement","UNIONBANK":"Union Bank of India",
    "UPL":"UPL","UNITDSPR":"United Spirits","VBL":"Varun Beverages","VEDL":"Vedanta",
    "VMM":"Vishal Mega Mart","VOLTAS":"Voltas","WAAREEENER":"Waaree Energies","WIPRO":"Wipro",
    "YESBANK":"Yes Bank","ETERNAL":"Eternal (Zomato)",
    "ATHERENERG":"Ather Energy","MAHABANK":"Bank of Maharashtra","SAGILITY":"Sagility India",
}

# NSE index symbols to exclude from F&O stock list
_FNO_EXCLUDE = {
    "NIFTY","BANKNIFTY","FINNIFTY","MIDCPNIFTY","NIFTYNXT50",
    "NIFTYIT","UNDERLYING","SENSEX","BANKEX","NIFTY50","NIFTY100",
    "NIFTYFPI","TMPV",
}

import json as _json, os as _os, datetime as _dt
_CACHE_FILE = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "fno_cache.json")

def _load_fno_cache():
    try:
        with open(_CACHE_FILE) as f:
            c = _json.load(f)
        syms  = c.get("symbols", [])
        names = c.get("names", {})
        age   = (_dt.date.today() - _dt.date.fromisoformat(c.get("updated", "2000-01-01"))).days
        if len(syms) >= 50:
            return syms, names, age
    except Exception:
        pass
    return None, {}, None

def _save_fno_cache(symbols, names):
    try:
        with open(_CACHE_FILE, "w") as f:
            _json.dump({"updated": str(_dt.date.today()), "symbols": symbols, "names": names}, f)
    except Exception:
        pass

@st.cache_data(ttl=86400, show_spinner=False)
def _fetch_fno_symbols():
    """
    Fetch live F&O symbol list (Zerodha) + company names (NSE equity master) weekly.
    Saves both to fno_cache.json. Falls back to cache, then hardcoded list.
    """
    cached_syms, cached_names, cache_age = _load_fno_cache()

    if cached_syms is None or cache_age >= 7:
        try:
            import requests, io
            # Step 1: F&O symbols from Zerodha
            r = requests.get("https://api.kite.trade/instruments", timeout=20)
            r.raise_for_status()
            df = pd.read_csv(io.StringIO(r.text))
            fno = df[(df["exchange"] == "NFO") & (df["instrument_type"] == "FUT")]
            syms = sorted({
                str(s).strip() for s in fno["name"].dropna().unique()
                if str(s).strip() not in _FNO_EXCLUDE
            })
            if len(syms) >= 50:
                # Step 2: Company names from NSE equity master (for new stocks not in hardcoded dict)
                names = dict(cached_names)
                try:
                    r2 = requests.get(
                        "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
                        timeout=20,
                    )
                    r2.raise_for_status()
                    eq = pd.read_csv(io.StringIO(r2.text))
                    sym_col  = eq.columns[0]
                    name_col = next((c for c in eq.columns if "NAME" in c.upper()), None)
                    if name_col:
                        for _, row in eq.iterrows():
                            s = str(row[sym_col]).strip()
                            n = str(row[name_col]).strip()
                            if s in syms and n and n.lower() != "nan":
                                names[s] = n.title()
                except Exception:
                    pass
                _save_fno_cache(syms, names)
                return syms, names, "Zerodha Live"
        except Exception:
            pass
        if cached_syms:
            return cached_syms, cached_names, f"Cache ({cache_age}d old)"
        return None, {}, "Bundled"

    return cached_syms, cached_names, "Zerodha Live"

_fno_live, _fno_cached_names, _fno_source = _fetch_fno_symbols()
# Hardcoded names take priority (cleaner); cached NSE names fill in new/unknown stocks
_COMPANY_NAMES_MERGED = {**_fno_cached_names, **_COMPANY_NAMES}
_FNO = (
    [(s, s, _COMPANY_NAMES_MERGED.get(s, s)) for s in _fno_live]
    if _fno_live
    else [(s, s, _COMPANY_NAMES.get(s, s)) for s in sorted(_COMPANY_NAMES)]
)

_SECTOR = {
    # Financial Services
    "ABCAPITAL":"Financial Services","AXISBANK":"Financial Services",
    "BAJAJFINSV":"Financial Services","BAJFINANCE":"Financial Services",
    "BAJAJHLDNG":"Financial Services","BANDHANBNK":"Financial Services",
    "BANKBARODA":"Financial Services","BANKINDIA":"Financial Services",
    "BSE":"Financial Services","CAMS":"Financial Services",
    "CANBK":"Financial Services","CDSL":"Financial Services",
    "CHOLAFIN":"Financial Services","HDFCAMC":"Financial Services",
    "HDFCBANK":"Financial Services","HDFCLIFE":"Financial Services",
    "ICICIBANK":"Financial Services","ICICIGI":"Financial Services",
    "ICICIPRULI":"Financial Services","IDFCFIRSTB":"Financial Services",
    "IEX":"Financial Services","INDUSINDBK":"Financial Services",
    "INDIANB":"Financial Services","IRFC":"Financial Services",
    "JIOFIN":"Financial Services","KFINTECH":"Financial Services",
    "KOTAKBANK":"Financial Services","LICHSGFIN":"Financial Services",
    "LICI":"Financial Services","LTF":"Financial Services",
    "MANAPPURAM":"Financial Services","MFSL":"Financial Services",
    "MOTILALOFS":"Financial Services","MUTHOOTFIN":"Financial Services",
    "NAM-INDIA":"Financial Services","PFC":"Financial Services",
    "POLICYBZR":"Financial Services","PNB":"Financial Services",
    "PNBHOUSING":"Financial Services","RBLBANK":"Financial Services",
    "RECLTD":"Financial Services","SBICARD":"Financial Services",
    "SBILIFE":"Financial Services","SHRIRAMFIN":"Financial Services",
    "SBIN":"Financial Services","UNIONBANK":"Financial Services",
    "YESBANK":"Financial Services","360ONE":"Financial Services",
    # Information Technology
    "HCLTECH":"IT","INFY":"IT","KPITTECH":"IT","LTM":"IT",
    "MPHASIS":"IT","OFSS":"IT","PERSISTENT":"IT","TECHM":"IT",
    "TCS":"IT","WIPRO":"IT","COFORGE":"IT","TATAELXSI":"IT",
    # Pharma
    "ALKEM":"Pharma","AUROPHARMA":"Pharma","BIOCON":"Pharma",
    "CIPLA":"Pharma","DIVISLAB":"Pharma","DRREDDY":"Pharma",
    "GLENMARK":"Pharma","LAURUSLABS":"Pharma","LUPIN":"Pharma",
    "MANKIND":"Pharma","SUNPHARMA":"Pharma","TORNTPHARM":"Pharma",
    "ZYDUSLIFE":"Pharma",
    # Automobile
    "ASHOKLEY":"Automobile","BAJAJ-AUTO":"Automobile","EICHERMOT":"Automobile",
    "FORCEMOT":"Automobile","HEROMOTOCO":"Automobile","M&M":"Automobile",
    "MARUTI":"Automobile","TATAMOTORS":"Automobile","TVSMOTOR":"Automobile",
    "UNOMINDA":"Automobile","MOTHERSON":"Automobile",
    # Capital Goods
    "ABB":"Capital Goods","BDL":"Capital Goods","BEL":"Capital Goods",
    "BHARATFORG":"Capital Goods","BLUESTARCO":"Capital Goods",
    "BOSCHLTD":"Capital Goods","CGPOWER":"Capital Goods",
    "CUMMINSIND":"Capital Goods","DIXON":"Capital Goods","HAL":"Capital Goods",
    "HAVELLS":"Capital Goods","KEI":"Capital Goods","POLYCAB":"Capital Goods",
    "SIEMENS":"Capital Goods","TIINDIA":"Capital Goods","VOLTAS":"Capital Goods",
    "KAYNES":"Capital Goods","PGEL":"Capital Goods","SONACOMS":"Capital Goods",
    "GVT&D":"Capital Goods","POWERINDIA":"Capital Goods","AMBER":"Capital Goods",
    "APLAPOLLO":"Capital Goods","SOLARINDS":"Capital Goods",
    # Chemicals
    "PIDILITIND":"Chemicals","PIIND":"Chemicals","SRF":"Chemicals","UPL":"Chemicals",
    # Cement
    "AMBUJACEM":"Cement","DALBHARAT":"Cement","SHREECEM":"Cement",
    "ULTRACEMCO":"Cement","GRASIM":"Cement",
    # Consumer Goods
    "BRITANNIA":"Consumer Goods","COLPAL":"Consumer Goods","DABUR":"Consumer Goods",
    "DMART":"Consumer Goods","GODREJCP":"Consumer Goods","GODFRYPHLP":"Consumer Goods",
    "ITC":"Consumer Goods","JUBLFOOD":"Consumer Goods","MARICO":"Consumer Goods",
    "NESTLEIND":"Consumer Goods","PAGEIND":"Consumer Goods","PATANJALI":"Consumer Goods",
    "RADICO":"Consumer Goods","TITAN":"Consumer Goods","TATACONSUM":"Consumer Goods",
    "UNITDSPR":"Consumer Goods","VBL":"Consumer Goods","KALYANKJIL":"Consumer Goods",
    "HINDUNILVR":"Consumer Goods","CROMPTON":"Consumer Goods",
    # Metals & Mining
    "HINDALCO":"Metals & Mining","HINDZINC":"Metals & Mining","JSWSTEEL":"Metals & Mining",
    "JINDALSTEL":"Metals & Mining","NATIONALUM":"Metals & Mining","NMDC":"Metals & Mining",
    "SAIL":"Metals & Mining","TATASTEEL":"Metals & Mining","VEDL":"Metals & Mining",
    # Oil & Gas
    "BPCL":"Oil & Gas","GAIL":"Oil & Gas","HINDPETRO":"Oil & Gas",
    "IOC":"Oil & Gas","OIL":"Oil & Gas","ONGC":"Oil & Gas","PETRONET":"Oil & Gas",
    # Real Estate
    "DLF":"Real Estate","GODREJPROP":"Real Estate","LODHA":"Real Estate",
    "OBEROIRLTY":"Real Estate","PHOENIXLTD":"Real Estate","PRESTIGE":"Real Estate",
    # Power
    "ADANIGREEN":"Power","ADANIPOWER":"Power","ADANIENSOL":"Power",
    "JSWENERGY":"Power","NHPC":"Power","NTPC":"Power","POWERGRID":"Power",
    "SUZLON":"Power","TATAPOWER":"Power","WAAREEENER":"Power",
    "PREMIERENE":"Power","INOXWIND":"Power",
    # Healthcare
    "APOLLOHOSP":"Healthcare","FORTIS":"Healthcare","MAXHEALTH":"Healthcare",
    # Infrastructure
    "ADANIENT":"Infrastructure","ADANIPORTS":"Infrastructure",
    "GMRAIRPORT":"Infrastructure","RVNL":"Infrastructure",
    "NBCC":"Infrastructure","IREDA":"Infrastructure","MAZDOCK":"Infrastructure",
    "COCHINSHIP":"Infrastructure",
    # Telecom
    "BHARTIARTL":"Telecom","IDEA":"Telecom","INDUSTOWER":"Telecom",
    # Logistics / Aviation
    "INDIGO":"Aviation","DELHIVERY":"Logistics","CONCOR":"Logistics",
    # Others
    "ETERNAL":"E-Commerce","NYKAA":"E-Commerce","PAYTM":"Fintech","SWIGGY":"E-Commerce",
    "ATHERENERG":"Automobile","MAHABANK":"Financial Services","SAGILITY":"Healthcare",
}

@st.cache_data(ttl=300)
def _fetch_fno_prices():
    import yfinance as yf
    nse_syms = [nse for _, nse, _ in _FNO]
    tickers  = [f"{s}.NS" for s in nse_syms]
    result = {}
    try:
        data = yf.download(tickers, period="2d", auto_adjust=True, progress=False)
        close_df = data["Close"]
        for nse, ticker in zip(nse_syms, tickers):
            try:
                series = close_df[ticker].dropna()
                if len(series) >= 2:
                    prev, cur = float(series.iloc[-2]), float(series.iloc[-1])
                    result[nse] = (round(cur, 2), round((cur - prev) / prev * 100, 2))
                elif len(series) == 1:
                    result[nse] = (round(float(series.iloc[-1]), 2), None)
            except Exception:
                pass
    except Exception:
        pass
    return result

# ── Stat bar ───────────────────────────────────────────────────────────────────
_s1, _s2, _s3 = st.columns(3)
_s1.metric("📋 F&O Stocks", len(_FNO))
_s2.markdown(f"""<div data-testid="metric-container" style="background:white;border-radius:10px;padding:1rem 1.2rem;box-shadow:0 1px 3px rgba(0,0,0,0.08);border-left:4px solid #3b82f6;">
<div style="color:#64748b;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:500;">🏦 F&O List Source</div>
<div style="color:#0f172a;font-size:0.85rem;font-weight:600;margin-top:4px;">{_fno_source}</div>
</div>""", unsafe_allow_html=True)
_s3.markdown("""<div data-testid="metric-container" style="background:white;border-radius:10px;padding:1rem 1.2rem;box-shadow:0 1px 3px rgba(0,0,0,0.08);border-left:4px solid #3b82f6;">
<div style="color:#64748b;font-size:0.72rem;text-transform:uppercase;letter-spacing:0.05em;font-weight:500;">🔄 Prices</div>
<div style="color:#0f172a;font-size:0.85rem;font-weight:600;margin-top:4px;">On demand via yfinance</div>
</div>""", unsafe_allow_html=True)

# ── Section 1: Probable Upside — Next 1–2 Weeks ──────────────────────────────
st.markdown(f"""
<div class="section-header">
    <span class="section-header-title">🔮 Probable Upside — Next 1–2 Weeks</span>
    <span class="section-badge">10 signals</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

st.caption("Screens all F&O stocks using 10 short-term technical signals — results cached for this session")

if st.button("🔮 Screen for Upside Candidates", key="load_swing"):
    st.session_state.pop("swing_data", None)
    st.session_state["swing_requested"] = True

if st.session_state.get("swing_requested"):
    if "swing_data" not in st.session_state:
        import yfinance as yf

        def _calc_ema(prices, n):
            k = 2 / (n + 1)
            e = prices[0]
            for p in prices[1:]:
                e = p * k + e * (1 - k)
            return e

        _nse_syms2 = [nse for _, nse, _ in _FNO]
        _tickers2  = [f"{s}.NS" for s in _nse_syms2]

        with st.spinner("Downloading price history for all F&O stocks…"):
            _hist2     = yf.download(_tickers2, period="1y", auto_adjust=True, progress=False)
            _close_df2 = _hist2["Close"]
            _vol_df2   = _hist2["Volume"]

        _swing_rows = []
        _UPSIDE_BLACKLIST = set()
        _prog2 = st.progress(0, text="Screening stocks…")

        for _i2, (_sym2, _nse2, _name2) in enumerate(_FNO):
            if _nse2 in _UPSIDE_BLACKLIST:
                continue
            _prog2.progress((_i2 + 1) / len(_FNO), text=f"Analysing {_nse2}…")
            try:
                _ticker2 = f"{_nse2}.NS"
                _cls_s   = _close_df2[_ticker2].dropna()
                _vol_s   = _vol_df2[_ticker2].reindex(_cls_s.index).fillna(0)
                _cls     = list(_cls_s.astype(float))
                _vols    = list(_vol_s.astype(float))
                _high_s  = _hist2["High"][_ticker2].reindex(_cls_s.index).fillna(method="ffill")
                _low_s   = _hist2["Low"][_ticker2].reindex(_cls_s.index).fillna(method="ffill")
                _highs   = list(_high_s.astype(float))
                _lows    = list(_low_s.astype(float))

                if len(_cls) < 60 or len(_highs) < 11 or len(_lows) < 11 or len(_vols) < 20:
                    continue

                _cur    = _cls[-1]
                _sma20  = sum(_cls[-20:]) / 20
                _sma50  = sum(_cls[-min(50,  len(_cls)):]) / min(50,  len(_cls))
                _sma200 = sum(_cls[-200:]) / 200 if len(_cls) >= 200 else sum(_cls) / len(_cls)
                _e20    = _calc_ema(_cls[-40:],  20) if len(_cls) >= 40  else _sma20
                _e50    = _calc_ema(_cls[-100:], 50) if len(_cls) >= 100 else _sma50

                # hard filter: above 200 SMA
                if _cur <= _sma200:
                    continue

                # RSI (14)
                _gains2  = [max(_cls[i] - _cls[i-1], 0) for i in range(1, len(_cls))]
                _losses2 = [max(_cls[i-1] - _cls[i], 0) for i in range(1, len(_cls))]
                _ag   = sum(_gains2[-14:])  / 14 if len(_gains2)  >= 14 else 0
                _al   = sum(_losses2[-14:]) / 14 if len(_losses2) >= 14 else 1
                _rsi2 = 100 if _al == 0 else 100 - (100 / (1 + _ag / _al))

                # hard filter: not overbought
                if _rsi2 >= 80:
                    continue

                _ema12    = _calc_ema(_cls[-30:], 12) if len(_cls) >= 30 else _cur
                _ema26    = _calc_ema(_cls[-50:], 26) if len(_cls) >= 50 else _cur
                _avgv20   = sum(_vols[-20:]) / 20

                # ── 10 setup signals ──────────────────────────────────────────
                # 1. Golden alignment: SMA20 > SMA50
                _golden_align  = _sma20 > _sma50

                # 2. Full trend: SMA50 > SMA200
                _sma_full      = _sma50 > _sma200

                # 3. SMA20 rising: recent 5D avg > prior 5D avg
                _sma20_rising  = (sum(_cls[-5:]) / 5) > (sum(_cls[-10:-5]) / 5)

                # 4. MACD positive: 12 EMA > 26 EMA
                _macd_pos      = _ema12 > _ema26

                # 5. RSI in healthy zone: not oversold, not overbought
                _rsi_zone      = 45 <= _rsi2 <= 72

                # 6. Net accumulation (10D): buying vol > 120% of selling vol
                _up_v10 = sum(_vols[i] for i in range(-10, 0) if _cls[i] >= _cls[i-1])
                _dn_v10 = sum(_vols[i] for i in range(-10, 0) if _cls[i] <  _cls[i-1])
                _net_accum = _up_v10 > _dn_v10 * 1.2

                # 7. Vol dry-up: today's vol < 70% of 20D avg (sellers absent)
                _vol_dryup = bool(_avgv20 and _vols[-1] < _avgv20 * 0.70)

                # 8. At support: price near 20 EMA or 50 EMA
                _at_support = (
                    (-0.02 <= (_cur / _e20 - 1) <= 0.04) or
                    (-0.01 <= (_cur / _e50 - 1) <= 0.03)
                )

                # 9. Rising lows (7D): floor trending up
                _rising_lows = bool(
                    len(_lows) >= 7 and
                    _lows[-1] > _lows[-4] and _lows[-4] > _lows[-7]
                )

                # 10. Entry trigger: vol expanding + closing up today
                _entry_trigger = bool(
                    len(_vols) >= 2 and len(_cls) >= 2 and
                    _vols[-1] > _vols[-2] and _cls[-1] > _cls[-2]
                )

                _st_score = sum([
                    _golden_align, _sma_full, _sma20_rising, _macd_pos,
                    _rsi_zone, _net_accum, _vol_dryup, _at_support,
                    _rising_lows, _entry_trigger,
                ])

                if _st_score == 10:
                    _quality = "Prime"
                elif _st_score >= 8:
                    _quality = "Sweet Spot"
                else:
                    _quality = "Strong"

                if _st_score >= 7:
                    _vol_ratio = round(_vols[-1] / _avgv20, 2) if _avgv20 else 0.0
                    _swing_rows.append({
                        "Stock":         _nse2,
                        "Sector":        _SECTOR.get(_nse2, "Other"),
                        "Price (₹)":     round(_cur, 2),
                        "RSI":           round(_rsi2, 1),
                        "Vol Ratio":     _vol_ratio,
                        "Net Accum":     "✅" if _net_accum      else "❌",
                        "At Support":    "✅" if _at_support     else "❌",
                        "Rising Lows":   "✅" if _rising_lows    else "❌",
                        "Entry Trigger": "✅" if _entry_trigger  else "❌",
                        "Full Trend":    "✅" if _sma_full       else "❌",
                        "Quality":       _quality,
                        "Score /10":     _st_score,
                        "Chart":         f"https://www.tradingview.com/chart/?symbol=NSE:{_nse2}",
                    })
            except Exception:
                pass

        _prog2.empty()
        st.session_state["swing_data"] = _swing_rows

    _swing_data = st.session_state.get("swing_data", [])
    if _swing_data:
        _SORT_PRIORITY = {"Prime": 0, "Sweet Spot": 1, "Strong": 2}
        _swing_df = pd.DataFrame(_swing_data)
        _swing_df["_pri"] = _swing_df["Quality"].map(_SORT_PRIORITY)
        _swing_df = (_swing_df
                     .sort_values(["_pri", "Score /10", "Vol Ratio"], ascending=[True, False, False])
                     .drop(columns=["_pri"])
                     .head(20)
                     .reset_index(drop=True))
        _swing_df.index += 1

        _prime = (_swing_df["Quality"] == "Prime").sum()
        _sweet = (_swing_df["Quality"] == "Sweet Spot").sum()
        st.markdown(
            f"**{len(_swing_df)} candidates** found (Score ≥ 7/10)  •  "
            f"**{_prime} prime** picks (Score 10/10)  •  "
            f"**{_sweet} sweet spot** (Score 8–9/10)"
        )

        st.dataframe(
            _swing_df,
            use_container_width=True,
            height=min(600, 56 + len(_swing_df) * 35),
            column_config={
                "Stock":          st.column_config.TextColumn("Stock"),
                "Sector":         st.column_config.TextColumn("Sector"),
                "Price (₹)":      st.column_config.NumberColumn("Price (₹)",     format="₹%.2f"),
                "RSI":            st.column_config.NumberColumn("RSI",            format="%.1f"),
                "Vol Ratio":      st.column_config.NumberColumn("Vol Ratio",      format="%.2fx"),
                "Net Accum":      st.column_config.TextColumn("Net Accum (10D)"),
                "At Support":     st.column_config.TextColumn("At Support"),
                "Rising Lows":    st.column_config.TextColumn("Rising Lows"),
                "Entry Trigger":  st.column_config.TextColumn("Entry Trigger"),
                "Full Trend":     st.column_config.TextColumn("SMA50>SMA200"),
                "Quality":        st.column_config.TextColumn("Quality"),
                "Score /10":      st.column_config.NumberColumn("Score /10",      format="%d ⭐"),
                "Chart":          st.column_config.LinkColumn("TradingView",      display_text="📈 Open Chart"),
            },
        )
        st.caption(
            "10 signals — Golden align · Full trend · SMA20 rising · MACD+ · RSI 45–72 · "
            "Net accum 10D · Vol dry-up · At support · Rising lows · Entry trigger  "
            "•  Hard filters: above SMA200 · RSI < 80  "
            "•  Quality: Prime (10/10) · Sweet Spot (8–9) · Strong (7)  "
            "•  ⚠️ For informational purposes only — not financial advice"
        )
    else:
        st.info("No stocks met the screening criteria. Try again during market hours.")

st.markdown("</div>", unsafe_allow_html=True)  # close section-body

# ── Section 2: Support Entry Screener ─────────────────────────────────────────
st.markdown("""
<div class="section-header">
    <span class="section-header-title">🎯 Support Entry Screener</span>
    <span class="section-badge">Buy at support</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

st.caption("Finds stocks that pulled back to a key support level and are showing early signs of reversal — potential low-risk buying opportunities")

if st.button("🎯 Find Support Entry Setups", key="load_support"):
    st.session_state.pop("support_data", None)
    st.session_state["support_requested"] = True

if st.session_state.get("support_requested"):
    if "support_data" not in st.session_state:
        import yfinance as yf

        _syms3 = [nse for _, nse, _ in _FNO]
        _tix3  = [f"{s}.NS" for s in _syms3]

        with st.spinner("Downloading price history for support analysis…"):
            _dh3  = yf.download(_tix3, period="1y", auto_adjust=True, progress=False)
            _cl3  = _dh3["Close"]
            _lo3  = _dh3["Low"]
            _hi3  = _dh3["High"]
            _vl3  = _dh3["Volume"]

        _sup_rows = []
        _prog3 = st.progress(0, text="Scanning support levels…")

        for _i3, (_, _nse3, _) in enumerate(_FNO):
            _prog3.progress((_i3 + 1) / len(_FNO), text=f"Scanning {_nse3}…")
            try:
                _tk3  = f"{_nse3}.NS"
                _cs3  = _cl3[_tk3].dropna()
                _c3   = list(_cs3.astype(float))
                _l3   = list(_lo3[_tk3].dropna().astype(float))
                _h3s  = list(_hi3[_tk3].dropna().astype(float))
                _v3   = list(_vl3[_tk3].reindex(_cs3.index).fillna(0).astype(float))

                if len(_c3) < 60 or len(_l3) < 10 or len(_h3s) < 10 or len(_v3) < 20:
                    continue

                _p3     = _c3[-1]
                _avgv3  = sum(_v3[-20:]) / 20 if len(_v3) >= 20 else 0

                # SMA levels
                _sm20_3  = sum(_c3[-20:]) / 20
                _sm50_3  = sum(_c3[-min(50, len(_c3)):]) / min(50, len(_c3))
                _sm200_3 = sum(_c3[-200:]) / 200 if len(_c3) >= 200 else sum(_c3) / len(_c3)

                # hard filter: above 200 SMA
                if _p3 <= _sm200_3:
                    continue

                # ── Swing lows in last 60 bars ─────────────────────────────
                _slows3 = []
                _st3 = max(3, len(_l3) - 60)
                for _si in range(_st3, len(_l3) - 3):
                    if (all(_l3[_si] <= _l3[_si - k] for k in range(1, 4) if _si - k >= 0)
                            and all(_l3[_si] <= _l3[_si + k] for k in range(1, 4))):
                        _slows3.append(_l3[_si])

                # ── Support candidates at or below current price ───────────
                _sups3 = [s for s in _slows3 if s <= _p3]
                for _sm3 in [_sm20_3, _sm50_3, _sm200_3]:
                    if _sm3 and _sm3 <= _p3:
                        _sups3.append(_sm3)
                if not _sups3:
                    continue

                _near3 = max(_sups3)   # closest support below price
                _gap3  = (_p3 - _near3) / _near3 * 100

                # hard filter: must be close to support
                if not (0 <= _gap3 <= 4.0):
                    continue

                # Pullback from 20D high
                _hi20_3 = max(_h3s[-20:]) if len(_h3s) >= 20 else max(_h3s)
                _pb3    = (_hi20_3 - _p3) / _hi20_3 * 100
                if not (3 <= _pb3 <= 25):
                    continue

                # RSI (Wilder's)
                _g3  = [max(_c3[i] - _c3[i-1], 0) for i in range(1, len(_c3))]
                _ls3 = [max(_c3[i-1] - _c3[i], 0) for i in range(1, len(_c3))]
                if len(_g3) >= 14:
                    _ag3 = sum(_g3[:14]) / 14
                    _al3 = sum(_ls3[:14]) / 14
                    for _gi, _lsi in zip(_g3[14:], _ls3[14:]):
                        _ag3 = (_ag3 * 13 + _gi) / 14
                        _al3 = (_al3 * 13 + _lsi) / 14
                    _rsi3 = 100 - (100 / (1 + _ag3 / max(_al3, 1e-10)))
                else:
                    _rsi3 = 50.0

                # ── 10 Base Formation Signals ──────────────────────────────

                # 1. Close to support (gap ≤ 1.5%)
                _at_sup3 = _gap3 <= 1.5

                # 2. Prior uptrend: SMA50 now >= SMA50 from 20 days ago
                _sma50_20ago_s   = sum(_c3[-70:-20]) / 50 if len(_c3) >= 70 else None
                _prior_uptrend_s = bool(_sma50_20ago_s and _sm50_3 >= _sma50_20ago_s)

                # 3. Confluence: 2+ support sources within 1.5% of nearest support
                _all_sups3   = [s for s in [_sm20_3, _sm50_3, _sm200_3] + _slows3 if s and s <= _p3]
                _confluence3 = sum(1 for s in _all_sups3 if abs(s - _near3) / _near3 <= 0.015) >= 2

                # 4. Support tested ≥ 2 times (proven level)
                _sup_touches, _in_zone = 0, False
                for _li in _l3[-120:]:
                    _at_sup_b = abs(_li - _near3) / _near3 <= 0.02
                    if _at_sup_b and not _in_zone:
                        _sup_touches += 1
                    _in_zone = _at_sup_b
                _proven_sup = _sup_touches >= 2

                # 5. Base formation: last 5D price range < 5% (tight consolidation)
                _base5_range = (max(_h3s[-5:]) - min(_l3[-5:])) / _p3 if len(_h3s) >= 5 else 1.0
                _base_forming = _base5_range < 0.05

                # 6. Volume condition: quiet base (declining vol) OR reversal surge (vol > 1.5x avg + up close)
                _vol5  = sum(_v3[-5:])  / 5  if len(_v3) >= 5  else _avgv3
                _vol10 = sum(_v3[-15:-5]) / 10 if len(_v3) >= 15 else _avgv3
                _vol_surge3 = bool(_avgv3 and _v3[-1] > _avgv3 * 1.5 and _c3[-1] > _c3[-2])
                _vol_declining = bool(_vol10 > 0 and _vol5 < _vol10) or _vol_surge3

                # 7. RSI reset: cooled to neutral zone (not in freefall, not overbought)
                _rsi_reset = 35 <= _rsi3 <= 65

                # 8. Lows stabilizing: closes not making new lows (use close to avoid penalising hammer wicks)
                _lows_stable = bool(len(_c3) >= 5 and _c3[-1] >= _c3[-5])

                # 9. Reversal candle: lower wick rejection OR 2-day consecutive up close
                _lower_wick3 = (
                    abs(_l3[-1] - _near3) / _near3 <= 0.02 and
                    (_c3[-1] - _l3[-1]) / max(_c3[-1], 1) > 0.005
                )
                _bounce3 = len(_c3) >= 3 and _c3[-1] > _c3[-2] and _c3[-2] > _c3[-3]
                _reversal = _lower_wick3 or _bounce3

                # 10. Entry trigger: vol expanding + up close (early sign) OR breaking above base top (strong sign)
                _base_top   = max(_c3[-6:-1]) if len(_c3) >= 6 else _c3[-2]
                _vol_exp3   = bool(len(_v3) >= 2 and _v3[-1] > _v3[-2] and _c3[-1] > _c3[-2])
                _break_up3  = bool(_c3[-1] > _base_top and _avgv3 and _v3[-1] > _avgv3 * 0.8)
                _entry3     = _vol_exp3 or _break_up3

                _sc3 = sum([
                    _at_sup3, _prior_uptrend_s, _confluence3, _proven_sup,
                    _base_forming, _vol_declining, _rsi_reset,
                    _lows_stable, _reversal, _entry3,
                ])

                if _sc3 >= 5:
                    _buy_lo3  = round(_near3 * 0.99, 2)
                    _buy_hi3  = round(_near3 * 1.02, 2)
                    _stop3    = round(_near3 * 0.97, 2)
                    # target: prior swing high (last 60D) or +8% whichever is higher
                    _swing_hi = max(_h3s[-60:]) if len(_h3s) >= 60 else max(_h3s)
                    _tgt3     = round(max(_swing_hi, _p3 * 1.08), 2)
                    _rwd3     = round((_tgt3 - _p3) / _p3 * 100, 1)
                    _rsk3     = round((_p3 - _stop3) / _p3 * 100, 1)
                    _rr3      = round(_rwd3 / _rsk3, 1) if _rsk3 > 0 else 0.0

                    # Support type label
                    if _sm200_3 and abs(_near3 - _sm200_3) / _sm200_3 < 0.012:
                        _stype3 = "SMA 200"
                    elif abs(_near3 - _sm50_3) / _sm50_3 < 0.012:
                        _stype3 = "SMA 50"
                    elif abs(_near3 - _sm20_3) / _sm20_3 < 0.012:
                        _stype3 = "SMA 20"
                    else:
                        _stype3 = "Swing Low"

                    _sup_rows.append({
                        "Stock":         _nse3,
                        "Sector":        _SECTOR.get(_nse3, "Other"),
                        "Price (₹)":     round(_p3, 2),
                        "Support (₹)":   round(_near3, 2),
                        "Support Type":  _stype3,
                        "Gap %":         round(_gap3, 2),
                        "Pullback %":    round(_pb3, 1),
                        "RSI":           round(_rsi3, 1),
                        "Base Range %":  round(_base5_range * 100, 1),
                        "Prior Trend":   "✅" if _prior_uptrend_s else "❌",
                        "Confluence":    "✅" if _confluence3     else "❌",
                        "Vol Decline":   "✅" if _vol_declining   else "❌",
                        "Lows Stable":   "✅" if _lows_stable     else "❌",
                        "Reversal":      "✅" if _reversal        else "❌",
                        "Entry Trigger": "✅" if _entry3          else "❌",
                        "Buy Zone":      f"₹{_buy_lo3}–{_buy_hi3}",
                        "Target (₹)":   _tgt3,
                        "Stop (₹)":     _stop3,
                        "R:R":          _rr3,
                        "Score /10":     _sc3,
                        "Chart":         f"https://www.tradingview.com/chart/?symbol=NSE:{_nse3}",
                    })
            except Exception:
                pass

        _prog3.empty()
        st.session_state["support_data"] = _sup_rows

    _sup_data = st.session_state.get("support_data", [])
    if _sup_data:
        _sup_df = (
            pd.DataFrame(_sup_data)
            .sort_values(["Score /10", "R:R", "Gap %"], ascending=[False, False, True])
            .head(20)
            .reset_index(drop=True)
        )
        _sup_df.index += 1

        _strong3 = (_sup_df["Score /10"] >= 8).sum()
        st.markdown(
            f"**{len(_sup_df)} stocks** found at key support levels  •  "
            f"**{_strong3} high-conviction setups** (Score ≥ 8/10)"
        )

        st.dataframe(
            _sup_df,
            use_container_width=True,
            height=min(600, 56 + len(_sup_df) * 35),
            column_config={
                "Stock":          st.column_config.TextColumn("Stock"),
                "Sector":         st.column_config.TextColumn("Sector"),
                "Price (₹)":      st.column_config.NumberColumn("Price (₹)",     format="₹%.2f"),
                "Support (₹)":    st.column_config.NumberColumn("Support (₹)",   format="₹%.2f"),
                "Support Type":   st.column_config.TextColumn("Support Type"),
                "Gap %":          st.column_config.NumberColumn("Gap %",          format="%.2f%%"),
                "Pullback %":     st.column_config.NumberColumn("Pullback %",     format="%.1f%%"),
                "RSI":            st.column_config.NumberColumn("RSI",            format="%.1f"),
                "Base Range %":   st.column_config.NumberColumn("Base Range 5D",  format="%.1f%%"),
                "Prior Trend":    st.column_config.TextColumn("Prior Trend"),
                "Confluence":     st.column_config.TextColumn("Confluence"),
                "Vol Decline":    st.column_config.TextColumn("Vol Condition"),
                "Lows Stable":    st.column_config.TextColumn("Lows Stable"),
                "Reversal":       st.column_config.TextColumn("Reversal"),
                "Entry Trigger":  st.column_config.TextColumn("Entry Trigger"),
                "Buy Zone":       st.column_config.TextColumn("Buy Zone"),
                "Target (₹)":    st.column_config.NumberColumn("Target (₹)",     format="₹%.2f"),
                "Stop (₹)":      st.column_config.NumberColumn("Stop (₹)",       format="₹%.2f"),
                "R:R":            st.column_config.NumberColumn("Reward:Risk",    format="%.1f:1"),
                "Score /10":      st.column_config.NumberColumn("Score /10",      format="%d ⭐"),
                "Chart":          st.column_config.LinkColumn("TradingView",      display_text="📈 Open Chart"),
            },
        )

        with st.expander("ℹ️ How the Support Entry Screener works"):
            st.markdown("""
| Signal | What it detects |
|---|---|
| **At Support (gap ≤ 1.5%)** | Price sitting right on support — optimal entry zone |
| **Prior Trend** | SMA50 now ≥ SMA50 from 20 days ago — pullback in an uptrend, not a breakdown |
| **Confluence** | 2+ support sources (swing low + SMA) within 1.5% — significantly stronger level |
| **Proven Support** | Level tested and held ≥ 2 times in past 120 days — not just theory |
| **Base Forming** | Last 5D price range < 5% — stock consolidating tightly, not falling |
| **Vol Condition** | Volume declining in base (sellers exhausting) OR volume surge today (>1.5× avg) with up close (reversal confirmed) |
| **RSI Reset (35–65)** | RSI cooled from overbought to neutral — healthy retracement |
| **Lows Stable** | Recent close ≥ close 5 days ago — trend floor holding (uses closes, not intraday lows, to avoid penalising hammer wicks) |
| **Reversal Candle** | Lower wick rejection at support OR 2 consecutive up-close days |
| **Entry Trigger** | Volume expanding + up close today (early sign) OR price breaking above 5D base top with normal+ volume (strong confirmation) |

**Target** = prior 60D swing high or +8% (whichever is higher) · **Stop** = 3% below support
⚠️ For informational purposes only — not financial advice
""")
    else:
        st.info("No stocks found at key support levels right now. Try during or after a broader market pullback.")

st.markdown("</div>", unsafe_allow_html=True)  # close section-body

# ── Section 3: Consolidation Breakout Screener ───────────────────────────────
st.markdown("""
<div class="section-header">
    <span class="section-header-title">🔲 Consolidation Breakout Screener</span>
    <span class="section-badge">Coiling to break out</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

st.caption("Finds stocks coiling in a tight range with shrinking volume and Bollinger squeeze — typically precede a strong upward breakout")

if st.button("🔲 Find Consolidation Setups", key="load_consol"):
    st.session_state.pop("consol_data", None)
    st.session_state["consol_requested"] = True

if st.session_state.get("consol_requested"):
    if "consol_data" not in st.session_state:
        import yfinance as yf

        _syms4 = [nse for _, nse, _ in _FNO]
        _tix4  = [f"{s}.NS" for s in _syms4]

        with st.spinner("Downloading price history for consolidation analysis…"):
            _dh4  = yf.download(_tix4, period="1y", auto_adjust=True, progress=False)
            _cl4  = _dh4["Close"]
            _hi4  = _dh4["High"]
            _vl4  = _dh4["Volume"]

        _con_rows = []
        _prog4 = st.progress(0, text="Scanning consolidation patterns…")

        for _i4, (_, _nse4, _) in enumerate(_FNO):
            _prog4.progress((_i4 + 1) / len(_FNO), text=f"Scanning {_nse4}…")
            try:
                _tk4  = f"{_nse4}.NS"
                _cs4  = _cl4[_tk4].dropna()
                _c4   = list(_cs4.astype(float))
                _h4   = list(_hi4[_tk4].dropna().astype(float))
                _v4   = list(_vl4[_tk4].reindex(_cs4.index).fillna(0).astype(float))

                if len(_c4) < 35:
                    continue

                _p4 = _c4[-1]

                # ── Range metrics ──────────────────────────────────────────
                _range10 = (max(_c4[-10:]) - min(_c4[-10:])) / _p4 * 100
                _range30 = (max(_c4[-30:]) - min(_c4[-30:])) / _p4 * 100

                # Hard filter: range must be tight (< 5%) and stock above SMA50
                _sma50_4_pre = sum(_c4[-min(50, len(_c4)):]) / min(50, len(_c4))
                if _range10 > 5 or _c4[-1] <= _sma50_4_pre:
                    continue
                # Range must be shrinking to less than half of 30D range
                _range_contract = _range10 < _range30 * 0.50

                # ── Days in consolidation: expand lookback until range exceeds 6% ──
                _days_consol = 10
                for _ext in range(11, min(60, len(_c4))):
                    _r_ext = (max(_c4[-_ext:]) - min(_c4[-_ext:])) / _p4 * 100
                    if _r_ext > 6:
                        break
                    _days_consol = _ext

                # ── Bollinger Band squeeze ─────────────────────────────────
                _sma20_4 = sum(_c4[-20:]) / 20
                _std20   = (sum((x - _sma20_4) ** 2 for x in _c4[-20:]) / 20) ** 0.5
                _bb_w    = 4 * _std20 / _sma20_4 * 100  # BB width %

                # Compare to BB width 15 days ago
                _sma20_p = sum(_c4[-35:-15]) / 20
                _std20_p = (sum((x - _sma20_p) ** 2 for x in _c4[-35:-15]) / 20) ** 0.5
                _bb_w_p  = 4 * _std20_p / _sma20_p * 100
                _bb_squeeze = _bb_w < _bb_w_p * 0.8  # current BB width < 80% of prior

                # ── Position in range ──────────────────────────────────────
                _high20_4 = max(_h4[-20:]) if len(_h4) >= 20 else max(_h4)
                _near_hi4 = _p4 >= _high20_4 * 0.95  # near top of range

                # ── Volume dry-up ──────────────────────────────────────────
                _avgv5_4  = sum(_v4[-5:])  / 5  if len(_v4) >= 5  else 0
                _avgv20_4 = sum(_v4[-20:]) / 20 if len(_v4) >= 20 else 0
                _vol_dry  = bool(_avgv20_4 and _avgv5_4 < _avgv20_4 * 0.85)
                _vol_ratio = round(_avgv5_4 / _avgv20_4, 2) if _avgv20_4 else 1.0

                # ── SMA20 flat ─────────────────────────────────────────────
                _sma20_5d = sum(_c4[-25:-5]) / 20 if len(_c4) >= 25 else _sma20_4
                _sma_flat = abs(_sma20_4 - _sma20_5d) / _sma20_4 < 0.012

                # ── Trend: price above SMA20 and SMA200 ───────────────────
                _sma50_4     = sum(_c4[-min(50, len(_c4)):]) / min(50, len(_c4))
                _sm200_4     = sum(_c4[-200:]) / 200 if len(_c4) >= 200 else None
                _above_sma20  = _p4 > _sma20_4
                _above_sma200 = bool(_sm200_4 and _p4 > _sm200_4)

                # ── Prior uptrend: SMA50 was rising 20 days before consolidation ──
                _sma50_20ago   = sum(_c4[-70:-20]) / 50 if len(_c4) >= 70 else None
                _prior_uptrend = bool(_sma50_20ago and _sma50_4 >= _sma50_20ago)

                # ── RSI (Wilder's EMA) ─────────────────────────────────────
                _g4  = [max(_c4[i] - _c4[i-1], 0) for i in range(1, len(_c4))]
                _ls4 = [max(_c4[i-1] - _c4[i], 0) for i in range(1, len(_c4))]
                if len(_g4) >= 14:
                    _ag4 = sum(_g4[:14]) / 14
                    _al4 = sum(_ls4[:14]) / 14
                    for _gi4, _lsi4 in zip(_g4[14:], _ls4[14:]):
                        _ag4 = (_ag4 * 13 + _gi4) / 14
                        _al4 = (_al4 * 13 + _lsi4) / 14
                    _rsi4 = 100 - (100 / (1 + _ag4 / max(_al4, 1e-10)))
                else:
                    _rsi4 = 50.0
                _rsi_ok4 = 45 <= _rsi4 <= 68

                # ── Breakout level = 0.5% above consolidation-period high ──
                _brk4       = round(max(_h4[-_days_consol:]) * 1.005, 2)
                _pct_to_brk = round((_brk4 - _p4) / _p4 * 100, 2)

                # ── Stop and Target (measured move) ───────────────────────
                _consol_lo4  = min(_c4[-_days_consol:])
                _consol_hi4  = max(_c4[-_days_consol:])
                _stop4       = round(_consol_lo4 * 0.98, 2)
                _rsk4        = round((_p4 - _stop4) / _p4 * 100, 1)
                _tgt4        = round(_brk4 + (_consol_hi4 - _consol_lo4), 2)

                # ── Score /10 ──────────────────────────────────────────────
                _sc4 = sum([
                    _range10 < 3.5,                 # very tight range (extra credit)
                    _range_contract,                 # range shrinking to < 50% of 30D range
                    _bb_squeeze,                     # Bollinger bands squeezing
                    _near_hi4,                       # consolidating near top of range
                    _vol_dry,                        # volume declining = accumulation
                    _sma_flat,                       # SMA20 flat = sideways action
                    _above_sma20,                    # price above SMA20
                    _above_sma200,                   # price above SMA200 = long-term uptrend
                    _prior_uptrend,                  # SMA50 was rising before consolidating
                    _rsi_ok4,                        # RSI 45-68: coiled, not extended
                ])

                if _sc4 >= 6:
                    _con_rows.append({
                        "Stock":        _nse4,
                        "Sector":       _SECTOR.get(_nse4, "Other"),
                        "Price (₹)":    round(_p4, 2),
                        "Breakout ₹":   _brk4,
                        "To Breakout%": _pct_to_brk,
                        "Target ₹":     _tgt4,
                        "Stop ₹":       _stop4,
                        "Risk %":       _rsk4,
                        "Days Consol.": _days_consol,
                        "10D Range %":  round(_range10, 2),
                        "BB Width %":   round(_bb_w, 2),
                        "Vol Ratio":    _vol_ratio,
                        "RSI":          round(_rsi4, 1),
                        "SMA Flat":     "✅" if _sma_flat       else "❌",
                        "Near High":    "✅" if _near_hi4       else "❌",
                        "BB Squeeze":   "✅" if _bb_squeeze     else "❌",
                        "Above SMA200": "✅" if _above_sma200   else "❌",
                        "Prior Trend":  "✅" if _prior_uptrend  else "❌",
                        "Score /10":    _sc4,
                        "Chart":        f"https://www.tradingview.com/chart/?symbol=NSE:{_nse4}",
                    })
            except Exception:
                pass

        _prog4.empty()
        st.session_state["consol_data"] = _con_rows

    _con_data = st.session_state.get("consol_data", [])
    if _con_data:
        _con_df = (
            pd.DataFrame(_con_data)
            .sort_values(["Score /10", "Days Consol.", "10D Range %"],
                         ascending=[False, False, True])
            .head(20)
            .reset_index(drop=True)
        )
        _con_df.index += 1

        _strong4 = (_con_df["Score /10"] >= 8).sum()
        st.markdown(
            f"**{len(_con_df)} stocks** in tight consolidation — watch for breakout above the Breakout ₹ level  •  "
            f"**{_strong4} high-conviction setups** (Score ≥ 8/10)"
        )

        st.dataframe(
            _con_df,
            use_container_width=True,
            height=min(600, 56 + len(_con_df) * 35),
            column_config={
                "Stock":        st.column_config.TextColumn("Stock"),
                "Sector":       st.column_config.TextColumn("Sector"),
                "Price (₹)":   st.column_config.NumberColumn("Price (₹)",      format="₹%.2f"),
                "Breakout ₹":  st.column_config.NumberColumn("Breakout ₹",     format="₹%.2f"),
                "To Breakout%":st.column_config.NumberColumn("To Breakout %",  format="%.2f%%"),
                "Target ₹":    st.column_config.NumberColumn("Target ₹",       format="₹%.2f"),
                "Stop ₹":      st.column_config.NumberColumn("Stop ₹",         format="₹%.2f"),
                "Risk %":      st.column_config.NumberColumn("Risk %",          format="%.1f%%"),
                "Days Consol.":st.column_config.NumberColumn("Days Consol.",   format="%d days"),
                "10D Range %": st.column_config.NumberColumn("10D Range %",    format="%.2f%%"),
                "BB Width %":  st.column_config.NumberColumn("BB Width %",     format="%.2f%%"),
                "Vol Ratio":   st.column_config.NumberColumn("Vol Ratio",      format="%.2fx"),
                "RSI":         st.column_config.NumberColumn("RSI",            format="%.1f"),
                "SMA Flat":    st.column_config.TextColumn("SMA Flat"),
                "Near High":   st.column_config.TextColumn("Near High"),
                "BB Squeeze":  st.column_config.TextColumn("BB Squeeze"),
                "Above SMA200":st.column_config.TextColumn("Above SMA200"),
                "Prior Trend": st.column_config.TextColumn("Prior Trend"),
                "Score /10":   st.column_config.NumberColumn("Score /10",      format="%d"),
                "Chart":       st.column_config.LinkColumn("TradingView",      display_text="📈 Open Chart"),
            },
        )

        with st.expander("ℹ️ How the Consolidation Breakout Screener works"):
            st.markdown("""
| Signal | What it detects |
|---|---|
| **10D Range %** | Price range of last 10 days as % of price — lower = tighter coil (filter: must be < 5%) |
| **Range Contraction** | 10-day range < 50% of 30-day range — the range is actively shrinking |
| **BB Squeeze** | Current Bollinger Band width < 80% of its level 15 days ago — volatility compressing |
| **Near High** | Price ≥ 95% of the 20-day high — consolidating at the top, not the bottom |
| **Vol Ratio** | 5-day avg volume ÷ 20-day avg volume — below 0.85x = volume drying up (accumulation) |
| **SMA Flat** | SMA20 moved < 1.2% over the past 5 days — confirms sideways / base-building action |
| **Trend Intact** | Price above SMA20 — near-term trend still up |
| **Above SMA200** | Price above SMA200 — long-term uptrend intact, not just a dead-cat bounce |
| **Prior Uptrend** | SMA50 now ≥ SMA50 from 20 days ago — stock was rising before it started consolidating |
| **RSI 45–68** | RSI in neutral-to-firm zone — coiled but not overbought |

**Score /10** — how many of the 10 signals are true. Min 6 required to appear.
**Sorted by:** Score (desc) → Days Consolidating (desc) → Range % (asc) — longest, tightest coils first.
**Entry signal:** Watch for a close above **Breakout ₹** on above-average volume.
**Target ₹** — measured move: breakout level + height of the consolidation range.
**Stop ₹** — 2% below the lowest close of the consolidation period.
**To Breakout %** — how far price needs to move to trigger; lower = closer to the edge.
⚠️ For informational purposes only — not financial advice.
""")
    else:
        st.info("No consolidation setups found right now. They appear after a period of trending followed by sideways action.")

st.markdown("</div>", unsafe_allow_html=True)  # close section-body

# ── Section 4: Backtest ────────────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
    <span class="section-header-title">📊 Backtest — Last 7 Days</span>
    <span class="section-badge">D+1 · D+3 · D+5 returns</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

st.caption(
    "Checks how each screener's picks performed over the next 1, 3, and 5 trading days. "
    "Reads the last 7 pick-dates from consolidated_history.xlsx."
)

if st.button("📊 Run Backtest", key="run_analysis_bt"):
    st.session_state.pop("analysis_bt_data", None)
    st.session_state["analysis_bt_requested"] = True

if st.session_state.get("analysis_bt_requested") and "analysis_bt_data" not in st.session_state:
    import yfinance as yf
    from openpyxl import load_workbook

    HIST_FILE = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "exports", "consolidated_history.xlsx")
    SHEETS    = ["Probable Upside", "Support Entry", "Consolidation Breakout"]

    if not _os.path.exists(HIST_FILE):
        st.warning("consolidated_history.xlsx not found. Run the daily export first.")
        st.session_state.pop("analysis_bt_requested", None)
    else:
        # ── Read wide-format history ───────────────────────────────────────────
        wb  = load_workbook(HIST_FILE, read_only=True, data_only=True)
        raw_picks = []   # (screener, date_str, stock)

        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws      = wb[sheet_name]
            rows    = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = rows[0]   # row 1 = date headers

            # collect last 7 unique trading dates for this sheet
            valid_dates = []
            for h in headers:
                if not h:
                    continue
                try:
                    d = _dt.date.fromisoformat(str(h)[:10])
                    if d.weekday() < 5:   # trading days only
                        valid_dates.append(str(d))
                except Exception:
                    pass
            last7 = sorted(set(valid_dates))[-7:]

            for ci, h in enumerate(headers):
                if not h:
                    continue
                try:
                    d_str = str(_dt.date.fromisoformat(str(h)[:10]))
                except Exception:
                    continue
                if d_str not in last7:
                    continue
                for row in rows[1:]:
                    if ci >= len(row) or not row[ci]:
                        continue
                    sym = str(row[ci]).strip()
                    if sym and sym.upper() != "NONE":
                        raw_picks.append({"Screener": sheet_name, "Date": d_str, "Stock": sym})

        wb.close()

        if not raw_picks:
            st.warning("No picks found in consolidated_history.xlsx.")
            st.session_state.pop("analysis_bt_requested", None)
        else:
            picks_df = pd.DataFrame(raw_picks)
            picks_df["Date"] = pd.to_datetime(picks_df["Date"])
            all_syms = picks_df["Stock"].unique().tolist()
            bt_tickers = [f"{s}.NS" for s in all_syms]

            with st.spinner(f"Downloading 3-month prices for {len(bt_tickers)} stocks…"):
                try:
                    dh = yf.download(bt_tickers, period="3mo", auto_adjust=True,
                                     progress=False, threads=True)
                    bt_close = {}
                    try:
                        sub = dh["Close"]
                    except KeyError:
                        sub = None
                    if sub is not None:
                        if isinstance(sub, pd.Series):
                            s = sub.dropna()
                            if len(s): bt_close[bt_tickers[0]] = s
                        else:
                            for col in sub.columns:
                                s = sub[col].dropna()
                                if len(s): bt_close[col] = s
                except Exception:
                    bt_close = {}

            today_ts  = pd.Timestamp.today().normalize()
            bt_rows   = []

            for _, row in picks_df.iterrows():
                pick_date = row["Date"]
                sym       = row["Stock"]
                tkr       = f"{sym}.NS"
                if tkr not in bt_close:
                    continue

                prices     = bt_close[tkr].sort_index()
                dates_list = list(prices.index)

                # last trading day on or before pick_date
                pick_pos = None
                for i, d in enumerate(dates_list):
                    if d <= pick_date:
                        pick_pos = i
                if pick_pos is None:
                    continue
                pick_px = float(prices.iloc[pick_pos])

                def _fwd(offset):
                    idx = pick_pos + offset
                    if idx >= len(dates_list) or dates_list[idx] > today_ts:
                        return None, None
                    p = float(prices.iloc[idx])
                    return round(p, 2), round((p - pick_px) / pick_px * 100, 2)

                p1, r1 = _fwd(1)
                p3, r3 = _fwd(3)
                p5, r5 = _fwd(5)

                bt_rows.append({
                    "Screener": row["Screener"],
                    "Date":     str(pick_date.date()),
                    "Stock":    sym,
                    "Pick ₹":   round(pick_px, 2),
                    "D+1 %":    r1,
                    "D+1":      "✅" if r1 and r1 > 0 else ("❌" if r1 is not None else "—"),
                    "D+3 %":    r3,
                    "D+3":      "✅" if r3 and r3 > 0 else ("❌" if r3 is not None else "—"),
                    "D+5 %":    r5,
                    "D+5":      "✅" if r5 and r5 > 0 else ("❌" if r5 is not None else "—"),
                })

            st.session_state["analysis_bt_data"] = bt_rows

if "analysis_bt_data" in st.session_state:
    bt_rows = st.session_state["analysis_bt_data"]

    if not bt_rows:
        st.info("Picks found but all forward prices are pending — check back after the next trading session.")
    else:
        bt_df = pd.DataFrame(bt_rows)

        def _hr(s):
            v = s.dropna()
            return f"{(v > 0).mean()*100:.1f}%" if len(v) else "N/A"
        def _avg(s):
            v = s.dropna()
            return f"{v.mean():.2f}%" if len(v) else "N/A"
        def _stats(df, label):
            r = {"Group": label, "Picks": len(df)}
            for h, col in [("1D","D+1 %"),("3D","D+3 %"),("5D","D+5 %")]:
                v = df[col].dropna()
                r[f"Hit Rate {h}"]   = f"{(v>0).mean()*100:.1f}%" if len(v) else "N/A"
                r[f"Avg Return {h}"] = f"{v.mean():.2f}%"          if len(v) else "N/A"
            return r

        # ── Overall summary ───────────────────────────────────────────────────
        st.markdown("#### Overall Summary")
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Total Picks",   len(bt_df))
        m2.metric("Dates Covered", bt_df["Date"].nunique())
        m3.metric("Hit Rate 1D",   _hr(bt_df["D+1 %"]))
        m4.metric("Hit Rate 3D",   _hr(bt_df["D+3 %"]))
        m5.metric("Hit Rate 5D",   _hr(bt_df["D+5 %"]))

        # ── Per-screener summary ──────────────────────────────────────────────
        st.markdown("#### By Screener")
        scr_cols = st.columns(3)
        for ci, screener in enumerate(["Probable Upside", "Support Entry", "Consolidation Breakout"]):
            sub = bt_df[bt_df["Screener"] == screener]
            if len(sub) == 0:
                continue
            scr_cols[ci].markdown(
                f"**{screener}** ({len(sub)} picks)  \n"
                f"Hit Rate — 1D: {_hr(sub['D+1 %'])} · 3D: {_hr(sub['D+3 %'])} · 5D: {_hr(sub['D+5 %'])}  \n"
                f"Avg Return — 1D: {_avg(sub['D+1 %'])} · 3D: {_avg(sub['D+3 %'])} · 5D: {_avg(sub['D+5 %'])}"
            )

        # ── Pick-by-pick table ─────────────────────────────────────────────────
        st.markdown("#### Pick-by-Pick Results")
        fc1, fc2, fc3 = st.columns(3)
        screener_filter = fc1.selectbox(
            "Filter by screener",
            ["All"] + ["Probable Upside", "Support Entry", "Consolidation Breakout"],
            key="bt_screener_filter"
        )
        all_dates = ["All"] + sorted(bt_df["Date"].unique().tolist(), reverse=True)
        date_filter = fc2.selectbox("Filter by date", all_dates, key="bt_date_filter")
        stock_filter = fc3.text_input("Search by stock", placeholder="e.g. RELIANCE", key="bt_stock_filter")

        disp = bt_df.copy()
        if screener_filter != "All":
            disp = disp[disp["Screener"] == screener_filter]
        if date_filter != "All":
            disp = disp[disp["Date"] == date_filter]
        if stock_filter.strip():
            disp = disp[disp["Stock"].str.contains(stock_filter.strip(), case=False, na=False)]

        st.dataframe(
            disp[["Screener","Date","Stock","Pick ₹","D+1 %","D+1","D+3 %","D+3","D+5 %","D+5"]],
            use_container_width=True,
            height=min(700, 56 + len(disp) * 35),
            column_config={
                "Pick ₹": st.column_config.NumberColumn("Pick Price", format="₹%.2f"),
                "D+1 %":  st.column_config.NumberColumn("D+1 Return", format="%+.2f%%"),
                "D+3 %":  st.column_config.NumberColumn("D+3 Return", format="%+.2f%%"),
                "D+5 %":  st.column_config.NumberColumn("D+5 Return", format="%+.2f%%"),
            },
            hide_index=True,
        )

        # ── Save Excel (upsert — update pending returns, append new picks) ────────
        bt_out = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), "exports", "backtest_results.xlsx")
        try:
            _RET_COLS = ["D+1 %", "D+1", "D+3 %", "D+3", "D+5 %", "D+5"]
            _KEY_COLS = ["Screener", "Date", "Stock"]

            def _upsert(existing, new):
                if existing is None or len(existing) == 0:
                    return new.copy()
                for df in [existing, new]:
                    for c in _KEY_COLS:
                        if c in df.columns:
                            df[c] = df[c].astype(str).str.strip()
                key_idx = {(r["Screener"], r["Date"], r["Stock"]): i
                           for i, r in existing.iterrows()}
                new_rows = []
                for _, row in new.iterrows():
                    k = (str(row["Screener"]), str(row["Date"]), str(row["Stock"]))
                    if k in key_idx:
                        ei = key_idx[k]
                        for col in _RET_COLS:
                            if col in row and col in existing.columns:
                                v = row[col]
                                if v is not None and str(v) not in ("—", "nan", "None", ""):
                                    existing.at[ei, col] = v
                    else:
                        new_rows.append(row.to_dict())
                if new_rows:
                    existing = pd.concat([existing, pd.DataFrame(new_rows)], ignore_index=True)
                return existing

            if _os.path.exists(bt_out):
                try:
                    existing_picks = pd.read_excel(bt_out, sheet_name="Pick Results", dtype=str)
                    combined = _upsert(existing_picks, bt_df.astype(str))
                except Exception:
                    combined = bt_df.astype(str)
            else:
                combined = bt_df.astype(str)

            for col in ["D+1 %", "D+3 %", "D+5 %"]:
                combined[col] = pd.to_numeric(combined[col], errors="coerce")

            # keep only last 30 days
            combined["Date"] = pd.to_datetime(combined["Date"], errors="coerce")
            cutoff = pd.Timestamp.today() - pd.Timedelta(days=30)
            combined = combined[combined["Date"] >= cutoff]
            combined["Date"] = combined["Date"].dt.strftime("%Y-%m-%d")

            summary_rows = [_stats(combined, "All Picks")]
            for screener in ["Probable Upside", "Support Entry", "Consolidation Breakout"]:
                sub = combined[combined["Screener"] == screener]
                if len(sub):
                    summary_rows.append(_stats(sub, screener))
            with pd.ExcelWriter(bt_out, engine="openpyxl") as writer:
                combined.to_excel(writer, sheet_name="Pick Results", index=False)
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        except Exception:
            pass
        st.caption(f"📁 Saved to exports/backtest_results.xlsx ({len(combined)} total rows)")

# ── Section 5: NSE F&O Stocks Table ──────────────────────────────────────────
st.markdown(f"""
<div class="section-header">
    <span class="section-header-title">📋 NSE F&amp;O Stocks</span>
    <span class="section-badge">{len(_FNO)} stocks</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

if st.button("📥 Load Today's Prices", key="load_fno_prices"):
    with st.spinner("Fetching today's prices…"):
        st.session_state["fno_prices"] = _fetch_fno_prices()

fno_df = pd.DataFrame([{
    "Stock":   nse_sym,
    "Company": name,
    "Chart":   f"https://www.tradingview.com/chart/?symbol=NSE:{nse_sym}",
} for _, nse_sym, name in _FNO])

_prices = st.session_state.get("fno_prices", {})
if _prices:
    fno_df["Price (₹)"] = fno_df["Stock"].map(lambda s: _prices.get(s, (None, None))[0])
    fno_df["Day (%)"]   = fno_df["Stock"].map(lambda s: _prices.get(s, (None, None))[1])

_fno_search = st.text_input(
    "🔍 Search by symbol or company",
    key="fno_search",
    placeholder="e.g. RELIANCE, Infosys, HDFC…",
)
if _fno_search.strip():
    _q = _fno_search.strip()
    _mask = (
        fno_df["Stock"].str.contains(_q, case=False, na=False)
        | fno_df["Company"].str.contains(_q, case=False, na=False)
    )
    fno_df = fno_df[_mask].reset_index(drop=True)

_col_cfg = {
    "Stock":   st.column_config.TextColumn("Stock"),
    "Company": st.column_config.TextColumn("Company"),
    "Chart":   st.column_config.LinkColumn("TradingView", display_text="📈 Open Chart"),
}
if _prices:
    _col_cfg["Price (₹)"] = st.column_config.NumberColumn("Price (₹)", format="₹%.2f")
    _col_cfg["Day (%)"]   = st.column_config.NumberColumn("Day (%)",   format="%.2f%%")

st.dataframe(
    fno_df,
    use_container_width=True,
    height=min(600, 56 + len(fno_df) * 35),
    column_config=_col_cfg,
    hide_index=True,
)
st.caption(f"Showing {len(fno_df)} of {len(_FNO)} NSE F&O stocks")
st.markdown("</div>", unsafe_allow_html=True)  # close section-body


st.markdown("</div>", unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)  # close content-wrap