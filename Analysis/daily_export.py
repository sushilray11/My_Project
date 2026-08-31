#!/usr/bin/env python3
"""
NSE F&O Screener — Daily Excel Export
Runs all 3 screeners and saves top-20 results to a dated .xlsx file.

Schedule via cron (3 PM IST, weekdays):
    crontab -e
    0 15 * * 1-5 cd /Users/I325211/Local_Project/Analysis && python3 daily_export.py >> exports/export_log.txt 2>&1
"""

import os, io, re, datetime, json
import pandas as pd
import yfinance as yf
import requests

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
EXPORTS_DIR = os.path.join(SCRIPT_DIR, "exports")
os.makedirs(EXPORTS_DIR, exist_ok=True)

TODAY     = datetime.date.today().strftime("%Y-%m-%d")
NOW       = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
OUTFILE   = os.path.join(EXPORTS_DIR, f"screener_{TODAY}.xlsx")
HIST_FILE = os.path.join(EXPORTS_DIR, "consolidated_history.xlsx")

# exit on weekends — no trading, no files
if datetime.date.today().weekday() >= 5:
    print(f"[{NOW}] Weekend — skipping export (market closed). Run on a weekday.", flush=True)
    raise SystemExit(0)

def log(msg):
    print(f"[{datetime.datetime.now():%H:%M:%S}] {msg}", flush=True)

# ── Company names ─────────────────────────────────────────────────────────────
COMPANY_NAMES = {
    "ABB":"ABB India","POWERINDIA":"Hitachi Energy India","ADANIENT":"Adani Enterprises",
    "ADANIGREEN":"Adani Green Energy","ADANIPORTS":"Adani Ports & SEZ","ADANIPOWER":"Adani Power",
    "ADANIENSOL":"Adani Energy Solutions","ABCAPITAL":"Aditya Birla Capital","ALKEM":"Alkem Laboratories",
    "GVT&D":"GE Vernova T&D India","AMBUJACEM":"Ambuja Cements","AMBER":"Amber Enterprises India",
    "ANGELONE":"Angel One","APLAPOLLO":"APL Apollo Tubes","APOLLOHOSP":"Apollo Hospitals",
    "ASHOKLEY":"Ashok Leyland","ASIANPAINT":"Asian Paints","ASTRAL":"Astral",
    "AUROPHARMA":"Aurobindo Pharma","AUBANK":"AU Small Finance Bank","DMART":"Avenue Supermarts",
    "AXISBANK":"Axis Bank","BAJAJ-AUTO":"Bajaj Auto","BAJAJFINSV":"Bajaj Finserv",
    "BAJFINANCE":"Bajaj Finance","BAJAJHLDNG":"Bajaj Holdings","BANDHANBNK":"Bandhan Bank",
    "BANKBARODA":"Bank of Baroda","BANKINDIA":"Bank of India","BHARTIARTL":"Bharti Airtel",
    "BDL":"Bharat Dynamics","BEL":"Bharat Electronics","BHARATFORG":"Bharat Forge",
    "INDUSTOWER":"Indus Towers","BPCL":"BPCL","BHEL":"BHEL","BIOCON":"Biocon",
    "BLUESTARCO":"Blue Star","BOSCHLTD":"Bosch India","BRITANNIA":"Britannia Industries",
    "BSE":"BSE","ZYDUSLIFE":"Zydus Lifesciences","CANBK":"Canara Bank","CDSL":"CDSL",
    "CHOLAFIN":"Cholamandalam Investment","CIPLA":"Cipla","COALINDIA":"Coal India",
    "COCHINSHIP":"Cochin Shipyard","COLPAL":"Colgate-Palmolive India","CAMS":"CAMS",
    "CONCOR":"Container Corp","CROMPTON":"Crompton Greaves Consumer","CGPOWER":"CG Power",
    "CUMMINSIND":"Cummins India","DABUR":"Dabur India","DELHIVERY":"Delhivery",
    "DIVISLAB":"Divi's Laboratories","DIXON":"Dixon Technologies","DLF":"DLF",
    "DRREDDY":"Dr. Reddy's Laboratories","EICHERMOT":"Eicher Motors","FEDERALBNK":"Federal Bank",
    "FORTIS":"Fortis Healthcare","FORCEMOT":"Force Motors","NYKAA":"Nykaa",
    "GAIL":"GAIL India","GLENMARK":"Glenmark Pharmaceuticals","GMRAIRPORT":"GMR Airports",
    "GODREJCP":"Godrej Consumer Products","GODFRYPHLP":"Godfrey Phillips India",
    "GODREJPROP":"Godrej Properties","GRASIM":"Grasim Industries","HAVELLS":"Havells India",
    "HCLTECH":"HCL Technologies","HDFCAMC":"HDFC AMC","HDFCBANK":"HDFC Bank",
    "HDFCLIFE":"HDFC Life Insurance","HEROMOTOCO":"Hero MotoCorp","HAL":"HAL",
    "HINDALCO":"Hindalco Industries","HINDUNILVR":"Hindustan Unilever","HINDPETRO":"HPCL",
    "HINDZINC":"Hindustan Zinc","HYUNDAI":"Hyundai Motor India","ICICIBANK":"ICICI Bank",
    "ICICIGI":"ICICI Lombard GI","ICICIPRULI":"ICICI Prudential Life","IDEA":"Vodafone Idea",
    "IDFCFIRSTB":"IDFC First Bank","360ONE":"360 ONE WAM","INDUSINDBK":"IndusInd Bank",
    "IEX":"Indian Energy Exchange","INDHOTEL":"Indian Hotels (Taj)","INDIANB":"Indian Bank",
    "IOC":"Indian Oil Corp","IRFC":"IRFC","IREDA":"IREDA","NAUKRI":"Info Edge (Naukri)",
    "INFY":"Infosys","INOXWIND":"INOX Wind","INDIGO":"IndiGo","ITC":"ITC",
    "JINDALSTEL":"Jindal Steel & Power","JIOFIN":"Jio Financial Services","JSWENERGY":"JSW Energy",
    "JSWSTEEL":"JSW Steel","JUBLFOOD":"Jubilant FoodWorks","KALYANKJIL":"Kalyan Jewellers",
    "KAYNES":"Kaynes Technology","KEI":"KEI Industries","KFINTECH":"KFin Technologies",
    "KOTAKBANK":"Kotak Mahindra Bank","KPITTECH":"KPIT Technologies","LT":"Larsen & Toubro",
    "LAURUSLABS":"Laurus Labs","LICI":"LIC India","LICHSGFIN":"LIC Housing Finance",
    "LTF":"L&T Finance","LTM":"LTIMindtree","LUPIN":"Lupin","LODHA":"Lodha Developers",
    "M&M":"Mahindra & Mahindra","MANAPPURAM":"Manappuram Finance","MANKIND":"Mankind Pharma",
    "MARICO":"Marico","MARUTI":"Maruti Suzuki","MFSL":"Max Financial Services",
    "MAXHEALTH":"Max Healthcare","MAZDOCK":"Mazagon Dock","MCX":"MCX",
    "UNOMINDA":"UNO Minda","MOTILALOFS":"Motilal Oswal","MOTHERSON":"Motherson Sumi",
    "MPHASIS":"Mphasis","MUTHOOTFIN":"Muthoot Finance","NATIONALUM":"National Aluminium",
    "NMDC":"NMDC","NBCC":"NBCC India","NESTLEIND":"Nestle India","NHPC":"NHPC",
    "COFORGE":"Coforge","NTPC":"NTPC","OBEROIRLTY":"Oberoi Realty","DALBHARAT":"Dalmia Bharat",
    "OIL":"Oil India","PAYTM":"Paytm","ONGC":"ONGC","OFSS":"Oracle Financial Services",
    "PAGEIND":"Page Industries","POLICYBZR":"PB Fintech","PERSISTENT":"Persistent Systems",
    "PETRONET":"Petronet LNG","PGEL":"PG Electroplast","PHOENIXLTD":"Phoenix Mills",
    "PIDILITIND":"Pidilite Industries","PIIND":"PI Industries","PNBHOUSING":"PNB Housing Finance",
    "POLYCAB":"Polycab India","PFC":"Power Finance Corp","POWERGRID":"Power Grid Corp",
    "PREMIERENE":"Premier Energies","PRESTIGE":"Prestige Estates","PNB":"Punjab National Bank",
    "RADICO":"Radico Khaitan","RVNL":"RVNL","RBLBANK":"RBL Bank","RELIANCE":"Reliance Industries",
    "NAM-INDIA":"Nippon India AMC","PATANJALI":"Patanjali Foods","RECLTD":"REC",
    "SAIL":"SAIL","SBICARD":"SBI Cards","SBILIFE":"SBI Life Insurance","SHREECEM":"Shree Cement",
    "SHRIRAMFIN":"Shriram Finance","SIEMENS":"Siemens India","SOLARINDS":"Solar Industries",
    "SONACOMS":"Sona BLW","SRF":"SRF","SBIN":"State Bank of India","SUNPHARMA":"Sun Pharma",
    "SUPREMEIND":"Supreme Industries","SUZLON":"Suzlon Energy","SWIGGY":"Swiggy",
    "TATAELXSI":"Tata Elxsi","TATACONSUM":"Tata Consumer Products","TATAMOTORS":"Tata Motors",
    "TATAPOWER":"Tata Power","TATASTEEL":"Tata Steel","TCS":"TCS","TECHM":"Tech Mahindra",
    "TITAN":"Titan Company","TORNTPHARM":"Torrent Pharmaceuticals","TRENT":"Trent",
    "TIINDIA":"Tube Investments","TVSMOTOR":"TVS Motor","ULTRACEMCO":"UltraTech Cement",
    "UNIONBANK":"Union Bank","UPL":"UPL","UNITDSPR":"United Spirits","VBL":"Varun Beverages",
    "VEDL":"Vedanta","VMM":"Vishal Mega Mart","VOLTAS":"Voltas","WAAREEENER":"Waaree Energies",
    "WIPRO":"Wipro","YESBANK":"Yes Bank","ETERNAL":"Eternal (Zomato)",
    "ATHERENERG":"Ather Energy","MAHABANK":"Bank of Maharashtra","SAGILITY":"Sagility India",
}

FNO_EXCLUDE = {
    "NIFTY","BANKNIFTY","FINNIFTY","MIDCPNIFTY","NIFTYNXT50",
    "NIFTYIT","UNDERLYING","SENSEX","BANKEX","NIFTY50","NIFTY100",
    "NIFTYFPI","TMPV",
}

CACHE_FILE = os.path.join(SCRIPT_DIR, "fno_cache.json")

def _load_fno_cache():
    try:
        with open(CACHE_FILE, encoding="utf-8") as f:
            c = json.load(f)
        syms  = c.get("symbols", [])
        names = c.get("names", {})
        age   = (datetime.date.today() - datetime.date.fromisoformat(c.get("updated", "2000-01-01"))).days
        if len(syms) >= 50:
            return syms, names, age
    except Exception:
        pass
    return None, {}, None

def _save_fno_cache(symbols, names):
    try:
        with open(CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump({"updated": str(datetime.date.today()), "symbols": symbols, "names": names}, f)
    except Exception:
        pass

# ── Helpers ───────────────────────────────────────────────────────────────────
def fetch_fno_symbols():
    cached_syms, cached_names, cache_age = _load_fno_cache()
    if cached_syms is None or cache_age >= 7:
        try:
            r = requests.get("https://api.kite.trade/instruments", timeout=20)
            r.raise_for_status()
            df = pd.read_csv(io.StringIO(r.text))
            fno = df[(df["exchange"] == "NFO") & (df["instrument_type"] == "FUT")]
            syms = sorted({
                str(s).strip() for s in fno["name"].dropna().unique()
                if str(s).strip() not in FNO_EXCLUDE
            })
            if len(syms) >= 50:
                # Also fetch company names from NSE equity master
                names = dict(cached_names)
                try:
                    r2 = requests.get(
                        "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
                        timeout=20,
                    )
                    r2.raise_for_status()
                    eq = pd.read_csv(io.StringIO(r2.text))
                    sym_col  = eq.columns[0]
                    name_col = next((c for c in eq.columns if "NAME" in c.upper()), None)
                    if name_col:
                        for _, row in eq.iterrows():
                            s = str(row[sym_col]).strip()
                            n = str(row[name_col]).strip()
                            if s in syms and n and n.lower() != "nan":
                                names[s] = n.title()
                except Exception:
                    pass
                _save_fno_cache(syms, names)
                return syms, names, "Zerodha Live"
        except Exception:
            pass
        if cached_syms:
            return cached_syms, cached_names, f"Cache ({cache_age}d old)"
        return None, {}, "Bundled"
    return cached_syms, cached_names, "Cache (fresh)"

SECTOR = {
    "ABCAPITAL":"Financial Services","AXISBANK":"Financial Services","BAJAJFINSV":"Financial Services",
    "BAJFINANCE":"Financial Services","BAJAJHLDNG":"Financial Services","BANDHANBNK":"Financial Services",
    "BANKBARODA":"Financial Services","BANKINDIA":"Financial Services","BSE":"Financial Services",
    "CAMS":"Financial Services","CANBK":"Financial Services","CDSL":"Financial Services",
    "CHOLAFIN":"Financial Services","HDFCAMC":"Financial Services","HDFCBANK":"Financial Services",
    "HDFCLIFE":"Financial Services","ICICIBANK":"Financial Services","ICICIGI":"Financial Services",
    "ICICIPRULI":"Financial Services","IDFCFIRSTB":"Financial Services","IEX":"Financial Services",
    "INDUSINDBK":"Financial Services","INDIANB":"Financial Services","IRFC":"Financial Services",
    "JIOFIN":"Financial Services","KFINTECH":"Financial Services","KOTAKBANK":"Financial Services",
    "LICHSGFIN":"Financial Services","LICI":"Financial Services","LTF":"Financial Services",
    "MANAPPURAM":"Financial Services","MFSL":"Financial Services","MOTILALOFS":"Financial Services",
    "MUTHOOTFIN":"Financial Services","NAM-INDIA":"Financial Services","PFC":"Financial Services",
    "POLICYBZR":"Financial Services","PNB":"Financial Services","PNBHOUSING":"Financial Services",
    "RBLBANK":"Financial Services","RECLTD":"Financial Services","SBICARD":"Financial Services",
    "SBILIFE":"Financial Services","SHRIRAMFIN":"Financial Services","SBIN":"Financial Services",
    "UNIONBANK":"Financial Services","YESBANK":"Financial Services","360ONE":"Financial Services",
    "HCLTECH":"IT","INFY":"IT","KPITTECH":"IT","LTM":"IT","MPHASIS":"IT","OFSS":"IT",
    "PERSISTENT":"IT","TECHM":"IT","TCS":"IT","WIPRO":"IT","COFORGE":"IT","TATAELXSI":"IT",
    "ALKEM":"Pharma","AUROPHARMA":"Pharma","BIOCON":"Pharma","CIPLA":"Pharma","DIVISLAB":"Pharma",
    "DRREDDY":"Pharma","GLENMARK":"Pharma","LAURUSLABS":"Pharma","LUPIN":"Pharma",
    "MANKIND":"Pharma","SUNPHARMA":"Pharma","TORNTPHARM":"Pharma","ZYDUSLIFE":"Pharma",
    "ASHOKLEY":"Automobile","BAJAJ-AUTO":"Automobile","EICHERMOT":"Automobile","FORCEMOT":"Automobile",
    "HEROMOTOCO":"Automobile","M&M":"Automobile","MARUTI":"Automobile","TATAMOTORS":"Automobile",
    "TVSMOTOR":"Automobile","UNOMINDA":"Automobile","MOTHERSON":"Automobile",
    "ABB":"Capital Goods","BDL":"Capital Goods","BEL":"Capital Goods","BHARATFORG":"Capital Goods",
    "BLUESTARCO":"Capital Goods","BOSCHLTD":"Capital Goods","CGPOWER":"Capital Goods",
    "CUMMINSIND":"Capital Goods","DIXON":"Capital Goods","HAL":"Capital Goods","HAVELLS":"Capital Goods",
    "KEI":"Capital Goods","POLYCAB":"Capital Goods","SIEMENS":"Capital Goods","TIINDIA":"Capital Goods",
    "VOLTAS":"Capital Goods","KAYNES":"Capital Goods","PGEL":"Capital Goods","SONACOMS":"Capital Goods",
    "GVT&D":"Capital Goods","POWERINDIA":"Capital Goods","AMBER":"Capital Goods",
    "APLAPOLLO":"Capital Goods","SOLARINDS":"Capital Goods",
    "PIDILITIND":"Chemicals","PIIND":"Chemicals","SRF":"Chemicals","UPL":"Chemicals",
    "AMBUJACEM":"Cement","DALBHARAT":"Cement","SHREECEM":"Cement","ULTRACEMCO":"Cement","GRASIM":"Cement",
    "BRITANNIA":"Consumer Goods","COLPAL":"Consumer Goods","DABUR":"Consumer Goods","DMART":"Consumer Goods",
    "GODREJCP":"Consumer Goods","GODFRYPHLP":"Consumer Goods","ITC":"Consumer Goods",
    "JUBLFOOD":"Consumer Goods","MARICO":"Consumer Goods","NESTLEIND":"Consumer Goods",
    "PAGEIND":"Consumer Goods","PATANJALI":"Consumer Goods","RADICO":"Consumer Goods",
    "TITAN":"Consumer Goods","TATACONSUM":"Consumer Goods","UNITDSPR":"Consumer Goods",
    "VBL":"Consumer Goods","KALYANKJIL":"Consumer Goods","HINDUNILVR":"Consumer Goods",
    "CROMPTON":"Consumer Goods",
    "HINDALCO":"Metals & Mining","HINDZINC":"Metals & Mining","JSWSTEEL":"Metals & Mining",
    "JINDALSTEL":"Metals & Mining","NATIONALUM":"Metals & Mining","NMDC":"Metals & Mining",
    "SAIL":"Metals & Mining","TATASTEEL":"Metals & Mining","VEDL":"Metals & Mining",
    "BPCL":"Oil & Gas","GAIL":"Oil & Gas","HINDPETRO":"Oil & Gas","IOC":"Oil & Gas",
    "OIL":"Oil & Gas","ONGC":"Oil & Gas","PETRONET":"Oil & Gas",
    "DLF":"Real Estate","GODREJPROP":"Real Estate","LODHA":"Real Estate","OBEROIRLTY":"Real Estate",
    "PHOENIXLTD":"Real Estate","PRESTIGE":"Real Estate",
    "ADANIGREEN":"Power","ADANIPOWER":"Power","ADANIENSOL":"Power","JSWENERGY":"Power",
    "NHPC":"Power","NTPC":"Power","POWERGRID":"Power","SUZLON":"Power","TATAPOWER":"Power",
    "WAAREEENER":"Power","PREMIERENE":"Power","INOXWIND":"Power",
    "APOLLOHOSP":"Healthcare","FORTIS":"Healthcare","MAXHEALTH":"Healthcare",
    "ADANIENT":"Infrastructure","ADANIPORTS":"Infrastructure","GMRAIRPORT":"Infrastructure",
    "RVNL":"Infrastructure","NBCC":"Infrastructure","IREDA":"Infrastructure",
    "MAZDOCK":"Infrastructure","COCHINSHIP":"Infrastructure",
    "BHARTIARTL":"Telecom","IDEA":"Telecom","INDUSTOWER":"Telecom",
    "INDIGO":"Aviation","DELHIVERY":"Logistics","CONCOR":"Logistics",
    "ETERNAL":"E-Commerce","NYKAA":"E-Commerce","PAYTM":"Fintech","SWIGGY":"E-Commerce",
    "ATHERENERG":"Automobile","MAHABANK":"Financial Services","SAGILITY":"Healthcare",
}

def calc_ema(prices, n):
    k = 2 / (n + 1)
    e = prices[0]
    for p in prices[1:]:
        e = p * k + e * (1 - k)
    return e

# ── Screener 1: Probable Upside ───────────────────────────────────────────────
_UPSIDE_BLACKLIST = set()

def run_probable_upside(fno, close_df, vol_df, low_df):
    rows = []
    for sym, nse, name in fno:
        if nse in _UPSIDE_BLACKLIST:
            continue
        try:
            ticker = f"{nse}.NS"
            cls_s  = close_df[ticker].dropna()
            vol_s  = vol_df[ticker].reindex(cls_s.index).fillna(0)
            low_s  = low_df[ticker].reindex(cls_s.index).ffill()
            cls    = list(cls_s.astype(float))
            vols   = list(vol_s.astype(float))
            lows   = list(low_s.astype(float))

            if len(cls) < 60 or len(lows) < 11 or len(vols) < 20:
                continue

            cur    = cls[-1]
            sma20  = sum(cls[-20:]) / 20
            sma50  = sum(cls[-min(50, len(cls)):]) / min(50, len(cls))
            sma200 = sum(cls[-200:]) / 200 if len(cls) >= 200 else sum(cls) / len(cls)

            # hard filter: above SMA200
            if cur <= sma200:
                continue

            # RSI (14)
            gains  = [max(cls[i] - cls[i-1], 0) for i in range(1, len(cls))]
            losses = [max(cls[i-1] - cls[i], 0) for i in range(1, len(cls))]
            ag = sum(gains[-14:])  / 14 if len(gains)  >= 14 else 0
            al = sum(losses[-14:]) / 14 if len(losses) >= 14 else 1
            rsi = 100 if al == 0 else 100 - (100 / (1 + ag / al))

            # hard filter: not overbought
            if rsi >= 80:
                continue

            ema12  = calc_ema(cls[-30:],  12) if len(cls) >= 30  else cur
            ema26  = calc_ema(cls[-50:],  26) if len(cls) >= 50  else cur
            e20    = calc_ema(cls[-40:],  20) if len(cls) >= 40  else sma20
            e50    = calc_ema(cls[-100:], 50) if len(cls) >= 100 else sma50
            avgv20 = sum(vols[-20:]) / 20

            # 1. Golden alignment: SMA20 > SMA50
            golden_align = sma20 > sma50
            # 2. Full trend: SMA50 > SMA200
            sma_full     = sma50 > sma200
            # 3. SMA20 rising
            sma20_rising = (sum(cls[-5:]) / 5) > (sum(cls[-10:-5]) / 5)
            # 4. MACD positive
            macd_pos     = ema12 > ema26
            # 5. RSI healthy zone
            rsi_zone     = 45 <= rsi <= 72
            # 6. Net accumulation (10D)
            up_v10 = sum(vols[i] for i in range(-10, 0) if cls[i] >= cls[i-1])
            dn_v10 = sum(vols[i] for i in range(-10, 0) if cls[i] <  cls[i-1])
            net_accum    = up_v10 > dn_v10 * 1.2
            # 7. Vol dry-up
            vol_dryup    = bool(avgv20 and vols[-1] < avgv20 * 0.70)
            # 8. At support (near EMA20 or EMA50)
            at_support   = (
                (-0.02 <= (cur / e20 - 1) <= 0.04) or
                (-0.01 <= (cur / e50 - 1) <= 0.03)
            )
            # 9. Rising lows (7D)
            rising_lows  = bool(
                len(lows) >= 7 and
                lows[-1] > lows[-4] and lows[-4] > lows[-7]
            )
            # 10. Entry trigger
            entry_trigger = bool(
                len(vols) >= 2 and len(cls) >= 2 and
                vols[-1] > vols[-2] and cls[-1] > cls[-2]
            )

            score = sum([
                golden_align, sma_full, sma20_rising, macd_pos,
                rsi_zone, net_accum, vol_dryup, at_support,
                rising_lows, entry_trigger,
            ])

            if score == 10:
                quality = "Prime"
            elif score >= 8:
                quality = "Sweet Spot"
            else:
                quality = "Strong"

            if score >= 7:
                vol_ratio = round(vols[-1] / avgv20, 2) if avgv20 else 0.0
                rows.append({
                    "Stock": nse, "Company": name,
                    "Sector": SECTOR.get(nse, "Other"),
                    "Price (₹)": round(cur, 2),
                    "RSI": round(rsi, 1),
                    "Vol Ratio": vol_ratio,
                    "Net Accum":    "Yes" if net_accum     else "No",
                    "At Support":   "Yes" if at_support    else "No",
                    "Rising Lows":  "Yes" if rising_lows   else "No",
                    "Entry Trigger":"Yes" if entry_trigger else "No",
                    "Full Trend":   "Yes" if sma_full      else "No",
                    "Quality":      quality,
                    "Score /10":    score,
                    "Chart": f"https://www.tradingview.com/chart/?symbol=NSE:{nse}",
                })
        except Exception:
            pass
    _SORT_PRIORITY = {"Prime": 0, "Sweet Spot": 1, "Strong": 2}
    df = pd.DataFrame(rows)
    if not df.empty:
        df["_pri"] = df["Quality"].map(_SORT_PRIORITY)
        df = (df.sort_values(["_pri", "Score /10", "Vol Ratio"], ascending=[True, False, False])
                .drop(columns=["_pri"])
                .head(20).reset_index(drop=True))
    df.index += 1
    return df

# ── Screener 2: Support Entry ─────────────────────────────────────────────────
def run_support_entry(fno, close_df, low_df, high_df, vol_df):
    rows = []
    for _, nse, _ in fno:
        try:
            tk  = f"{nse}.NS"
            cs  = close_df[tk].dropna()
            c   = list(cs.astype(float))
            l   = list(low_df[tk].dropna().astype(float))
            hs  = list(high_df[tk].dropna().astype(float))
            v   = list(vol_df[tk].reindex(cs.index).fillna(0).astype(float))
            if len(c) < 60 or len(l) < 10 or len(hs) < 10 or len(v) < 20:
                continue
            p    = c[-1]
            avgv = sum(v[-20:]) / 20 if len(v) >= 20 else 0

            sm20  = sum(c[-20:]) / 20
            sm50  = sum(c[-min(50, len(c)):]) / min(50, len(c))
            sm200 = sum(c[-200:]) / 200 if len(c) >= 200 else sum(c) / len(c)

            # hard filter: above SMA200
            if p <= sm200:
                continue

            # swing lows: 5-bar pivot, 90-bar lookback
            slows = []
            st = max(5, len(l) - 90)
            for si in range(st, len(l) - 5):
                if (all(l[si] <= l[si - k] for k in range(1, 6) if si - k >= 0)
                        and all(l[si] <= l[si + k] for k in range(1, 6))):
                    slows.append(l[si])

            sups = [s for s in slows if s <= p]
            for sm in [sm20, sm50, sm200]:
                if sm and sm <= p:
                    sups.append(sm)
            if not sups:
                continue

            near = max(sups)
            gap  = (p - near) / near * 100

            # hard filter: within 2.5% of support
            if not (0 <= gap <= 2.5):
                continue

            # hard filter: pullback 3–25% from 20D high
            hi20 = max(hs[-20:]) if len(hs) >= 20 else max(hs)
            pb   = (hi20 - p) / hi20 * 100
            if not (3 <= pb <= 25):
                continue

            # RSI (Wilder's)
            g  = [max(c[i] - c[i-1], 0) for i in range(1, len(c))]
            ls = [max(c[i-1] - c[i], 0) for i in range(1, len(c))]
            if len(g) >= 14:
                ag = sum(g[:14]) / 14
                al = sum(ls[:14]) / 14
                for gi, li_r in zip(g[14:], ls[14:]):
                    ag = (ag * 13 + gi) / 14
                    al = (al * 13 + li_r) / 14
                rsi = 100 - (100 / (1 + ag / max(al, 1e-10)))
            else:
                rsi = 50.0

            # 1. At support (gap ≤ 1.5%)
            at_sup = gap <= 1.5

            # 2. Prior uptrend
            sma50_20ago = sum(c[-70:-20]) / 50 if len(c) >= 70 else None
            prior_up = bool(sma50_20ago and sm50 >= sma50_20ago)

            # 3. Confluence: 2+ sources within 1.5%
            all_sups   = [s for s in [sm20, sm50, sm200] + slows if s and s <= p]
            confluence = sum(1 for s in all_sups if abs(s - near) / near <= 0.015) >= 2

            # 4. Proven support with recency weight (last 30 bars count double)
            sup_touches, in_zone = 0, False
            for idx, li in enumerate(l[-120:]):
                at_sup_b = abs(li - near) / near <= 0.02
                if at_sup_b and not in_zone:
                    sup_touches += 2 if idx >= 90 else 1
                in_zone = at_sup_b
            proven_sup = sup_touches >= 2

            # 5. Base forming: 5D range < 5%
            base5_range  = (max(hs[-5:]) - min(l[-5:])) / p if len(hs) >= 5 else 1.0
            base_forming = base5_range < 0.05

            # 6. Vol condition: split into quiet and surge
            vol5  = sum(v[-5:])  / 5  if len(v) >= 5  else avgv
            vol10 = sum(v[-15:-5]) / 10 if len(v) >= 15 else avgv
            vol_quiet = bool(vol10 > 0 and vol5 < vol10)
            vol_surge = bool(avgv and v[-1] > avgv * 1.5 and c[-1] > c[-2])
            vol_cond  = vol_quiet or vol_surge

            # 7. RSI reset 35–70
            rsi_reset = 35 <= rsi <= 70

            # 8. Lows stable
            lows_stable = bool(len(c) >= 5 and c[-1] >= c[-5])

            # 9. Reversal candle
            lower_wick = (
                abs(l[-1] - near) / near <= 0.02 and
                (c[-1] - l[-1]) / max(c[-1], 1) > 0.005
            )
            bounce  = len(c) >= 3 and c[-1] > c[-2] and c[-2] > c[-3]
            reversal = lower_wick or bounce

            # 10. Entry trigger
            base_top = max(c[-6:-1]) if len(c) >= 6 else c[-2]
            vol_exp  = bool(len(v) >= 2 and v[-1] > v[-2] and c[-1] > c[-2])
            break_up = bool(c[-1] > base_top and avgv and v[-1] > avgv * 0.8)
            entry    = vol_exp or break_up

            score = sum([at_sup, prior_up, confluence, proven_sup,
                         base_forming, vol_cond, rsi_reset,
                         lows_stable, reversal, entry])

            if score >= 5:
                buy_lo   = round(near * 0.99, 2)
                buy_hi   = round(near * 1.02, 2)
                stop     = round(near * 0.97, 2)
                swing_hi = max(hs[-60:]) if len(hs) >= 60 else max(hs)
                tgt      = round(max(swing_hi, p * 1.08), 2)
                rwd      = round((tgt - p) / p * 100, 1)
                rsk      = round((p - stop) / p * 100, 1)
                rr       = round(rwd / rsk, 1) if rsk > 0 else 0.0

                # hard filter: R:R ≥ 1.5
                if rr < 1.5:
                    continue

                if sm200 and abs(near - sm200) / sm200 < 0.012:
                    stype = "SMA 200"
                elif abs(near - sm50) / sm50 < 0.012:
                    stype = "SMA 50"
                elif abs(near - sm20) / sm20 < 0.012:
                    stype = "SMA 20"
                else:
                    stype = "Swing Low"

                rows.append({
                    "Stock": nse, "Company": COMPANY_NAMES.get(nse, nse),
                    "Sector": SECTOR.get(nse, "Other"),
                    "Price (₹)": round(p, 2), "Support (₹)": round(near, 2),
                    "Support Type": stype,
                    "Gap %": round(gap, 2), "Pullback %": round(pb, 1),
                    "RSI": round(rsi, 1),
                    "Base Range %": round(base5_range * 100, 1),
                    "Prior Trend": "Yes" if prior_up    else "No",
                    "Confluence":  "Yes" if confluence  else "No",
                    "Vol Quiet":   "Yes" if vol_quiet   else "No",
                    "Vol Surge":   "Yes" if vol_surge   else "No",
                    "Lows Stable": "Yes" if lows_stable else "No",
                    "Reversal":    "Yes" if reversal    else "No",
                    "Entry Trigger":"Yes" if entry      else "No",
                    "Buy Zone": f"₹{buy_lo}–{buy_hi}",
                    "Target (₹)": tgt, "Stop (₹)": stop,
                    "R:R": rr, "Score /10": score,
                    "Chart": f"https://www.tradingview.com/chart/?symbol=NSE:{nse}",
                })
        except Exception:
            pass
    df = (pd.DataFrame(rows)
          .sort_values(["Score /10", "R:R", "Gap %"], ascending=[False, False, True])
          .head(20).reset_index(drop=True))
    df.index += 1
    return df

# ── Screener 3: Consolidation Breakout ────────────────────────────────────────
def run_consolidation(fno, close_df, high_df, vol_df):
    rows = []
    for _, nse, _ in fno:
        try:
            tk  = f"{nse}.NS"
            cs  = close_df[tk].dropna()
            c   = list(cs.astype(float))
            h   = list(high_df[tk].dropna().astype(float))
            v   = list(vol_df[tk].reindex(cs.index).fillna(0).astype(float))
            if len(c) < 35:
                continue
            p = c[-1]
            range10   = (max(c[-10:]) - min(c[-10:])) / p * 100
            range30   = (max(c[-30:]) - min(c[-30:])) / p * 100
            sma50_pre = sum(c[-min(50, len(c)):]) / min(50, len(c))

            # hard filter: range < 4% and above SMA50
            if range10 > 4 or c[-1] <= sma50_pre:
                continue
            range_contract = range10 < range30 * 0.50

            # days in consolidation
            days_consol = 10
            for ext in range(11, min(60, len(c))):
                r_ext = (max(c[-ext:]) - min(c[-ext:])) / p * 100
                if r_ext > 6:
                    break
                days_consol = ext

            # hard filter: min 15 days
            if days_consol < 15:
                continue

            # prior move strength: ≥ 8% in 40D before consolidation
            pre_start  = max(0, len(c) - days_consol - 1)
            pre_40     = max(0, pre_start - 40)
            prior_move = ((c[pre_start] - c[pre_40]) / c[pre_40] * 100
                          if c[pre_40] > 0 else 0)
            strong_move = prior_move >= 8.0

            # vol pattern: high vol before → quiet now
            pre_vol5   = (sum(v[max(0, pre_start - 5):pre_start]) / 5
                          if pre_start >= 5 else 0)
            avgv5_now  = sum(v[-5:]) / 5 if len(v) >= 5 else 0
            vol_pattern = bool(pre_vol5 > 0 and avgv5_now > 0
                               and pre_vol5 >= avgv5_now * 1.2)

            # Bollinger Band squeeze
            sma20   = sum(c[-20:]) / 20
            std20   = (sum((x - sma20) ** 2 for x in c[-20:]) / 20) ** 0.5
            bb_w    = 4 * std20 / sma20 * 100
            sma20_p = sum(c[-35:-15]) / 20
            std20_p = (sum((x - sma20_p) ** 2 for x in c[-35:-15]) / 20) ** 0.5
            bb_w_p  = 4 * std20_p / sma20_p * 100
            bb_sq   = bb_w < bb_w_p * 0.8

            avgv5    = sum(v[-5:])  / 5  if len(v) >= 5  else 0
            avgv20   = sum(v[-20:]) / 20 if len(v) >= 20 else 0
            vol_dry  = bool(avgv20 and avgv5 < avgv20 * 0.85)
            vol_ratio = round(avgv5 / avgv20, 2) if avgv20 else 1.0

            sma20_5d = sum(c[-25:-5]) / 20 if len(c) >= 25 else sma20
            sma_flat = abs(sma20 - sma20_5d) / sma20 < 0.012

            sm200        = sum(c[-200:]) / 200 if len(c) >= 200 else None
            above_sma200 = bool(sm200 and p > sm200)

            sma50_20ago = sum(c[-70:-20]) / 50 if len(c) >= 70 else None
            prior_up    = bool(sma50_20ago and sma50_pre >= sma50_20ago)

            # RSI (Wilder's)
            gains4  = [max(c[i] - c[i-1], 0) for i in range(1, len(c))]
            losses4 = [max(c[i-1] - c[i], 0) for i in range(1, len(c))]
            if len(gains4) >= 14:
                ag4 = sum(gains4[:14]) / 14
                al4 = sum(losses4[:14]) / 14
                for gi4, li4 in zip(gains4[14:], losses4[14:]):
                    ag4 = (ag4 * 13 + gi4) / 14
                    al4 = (al4 * 13 + li4) / 14
                rsi4 = 100 - (100 / (1 + ag4 / max(al4, 1e-10)))
            else:
                rsi4 = 50.0
            rsi_ok4 = 45 <= rsi4 <= 68

            brk        = round(max(h[-days_consol:]) * 1.005, 2)
            pct_to_brk = round((brk - p) / p * 100, 2)
            near_brk   = pct_to_brk <= 1.5

            consol_lo = min(c[-days_consol:])
            consol_hi = max(c[-days_consol:])
            stop4     = round(consol_lo * 0.98, 2)
            rsk4      = round((p - stop4) / p * 100, 1)
            tgt4      = round(brk + (consol_hi - consol_lo), 2)

            score = sum([
                strong_move, vol_pattern, range_contract, bb_sq,
                near_brk, vol_dry, sma_flat, above_sma200,
                prior_up, rsi_ok4,
            ])

            if score >= 7:
                rows.append({
                    "Stock": nse, "Company": COMPANY_NAMES.get(nse, nse),
                    "Sector": SECTOR.get(nse, "Other"),
                    "Price (₹)": round(p, 2), "Breakout ₹": brk,
                    "To Breakout%": pct_to_brk, "Target ₹": tgt4,
                    "Stop ₹": stop4, "Risk %": rsk4,
                    "Days Consol.": days_consol,
                    "10D Range %": round(range10, 2), "BB Width %": round(bb_w, 2),
                    "Vol Ratio": vol_ratio, "RSI": round(rsi4, 1),
                    "SMA Flat":     "Yes" if sma_flat     else "No",
                    "Prior Move":   "Yes" if strong_move  else "No",
                    "Vol Pattern":  "Yes" if vol_pattern  else "No",
                    "Near Breakout":"Yes" if near_brk     else "No",
                    "BB Squeeze":   "Yes" if bb_sq        else "No",
                    "Above SMA200": "Yes" if above_sma200 else "No",
                    "Prior Trend":  "Yes" if prior_up     else "No",
                    "Score /10": score,
                    "Chart": f"https://www.tradingview.com/chart/?symbol=NSE:{nse}",
                })
        except Exception:
            pass
    df = (pd.DataFrame(rows)
          .sort_values(["Score /10", "Days Consol.", "10D Range %"],
                       ascending=[False, False, True])
          .head(20).reset_index(drop=True))
    df.index += 1
    return df

# ── Excel export with formatting ──────────────────────────────────────────────
def write_excel(sheets_data):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    HEADER_FILL  = PatternFill("solid", fgColor="1D4ED8")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    GREEN_FILL   = PatternFill("solid", fgColor="D1FAE5")
    YELLOW_FILL  = PatternFill("solid", fgColor="FEF9C3")
    ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")
    THIN         = Side(style="thin", color="E2E8F0")
    CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    SCORE_COLS   = {"Score /11", "Score /10", "Score /9", "Score /8", "Score /6"}

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    tab_colors = {"Probable Upside": "3B82F6", "Support Entry": "10B981",
                  "Consolidation Breakout": "8B5CF6"}

    for sheet_name, df in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_properties.tabColor = tab_colors.get(sheet_name, "3B82F6")

        if df.empty:
            ws["A1"] = "No results found for this screener."
            continue

        cols = list(df.columns)

        # Header row
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill   = HEADER_FILL
            cell.font   = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = CELL_BORDER

        ws.row_dimensions[1].height = 28

        # Data rows
        score_col_idx = next((ci + 1 for ci, c in enumerate(cols) if c in SCORE_COLS), None)
        max_score = df[next((c for c in cols if c in SCORE_COLS), cols[-1])].max() if score_col_idx else 0

        for ri, (_, row) in enumerate(df.iterrows(), start=2):
            row_fill = ALT_FILL if ri % 2 == 0 else None
            for ci, col in enumerate(cols, start=1):
                val  = row[col]
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = CELL_BORDER
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if row_fill:
                    cell.fill = row_fill
                # Color score column
                if score_col_idx and ci == score_col_idx and isinstance(val, (int, float)):
                    cell.fill = GREEN_FILL if val >= max_score - 1 else YELLOW_FILL
                    cell.font = Font(bold=True, size=10)

        # Auto column width
        for ci, col in enumerate(cols, start=1):
            max_len = len(str(col))
            for ri in range(2, ws.max_row + 1):
                v = ws.cell(row=ri, column=ci).value
                max_len = max(max_len, len(str(v)) if v is not None else 0)
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 3, 30)

        ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.sheet_properties.tabColor = "0F172A"
    summary_rows = [
        ["NSE F&O Screener — Daily Export"],
        ["Generated", NOW],
        [""],
    ]
    for name, df in sheets_data.items():
        summary_rows.append([name, f"{len(df)} stocks found"])
    for ri, row in enumerate(summary_rows, start=1):
        for ci, val in enumerate(row, start=1):
            cell = ws_sum.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = Font(bold=True, size=13, color="1D4ED8")
            elif ri >= 4:
                cell.font = Font(size=10)
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    wb.save(OUTFILE)

# ── Consolidated history ──────────────────────────────────────────────────────
def update_consolidated_history(results: dict):
    """
    Layout per sheet:
      Row 1 : date1  | date2  | date3  | …  (oldest → newest, left → right)
      Row 2+: rank-1 stock | rank-2 stock | …  for that date
    Keeps 1 year of date columns.
    """
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # use last market-open day so weekend runs don't create non-trading dates
    _d = datetime.date.today()
    while _d.weekday() >= 5:
        _d -= datetime.timedelta(days=1)
    today_str = str(_d)
    cutoff    = datetime.date.today() - datetime.timedelta(days=365)

    SCORE_COL = {
        "Probable Upside":        "Score /10",
        "Support Entry":          "Score /10",
        "Consolidation Breakout": "Score /10",
    }

    # history[sheet][date_str] = [stock1, stock2, …]  (ranked, top first)
    history = {name: {} for name in results}

    # ── Load existing (new format: row1=dates, rows2+=stocks) ─────────────────
    if os.path.exists(HIST_FILE):
        try:
            wb_old = load_workbook(HIST_FILE)
            for sheet_name in results:
                if sheet_name not in wb_old.sheetnames:
                    continue
                ws = wb_old[sheet_name]
                headers = [cell.value for cell in ws[1]]
                # Only load if first header looks like a date (new format)
                try:
                    datetime.date.fromisoformat(str(headers[0])[:10])
                except Exception:
                    continue   # old format — skip, rebuild fresh
                for ci, h in enumerate(headers):
                    if not h:
                        continue
                    try:
                        d = datetime.date.fromisoformat(str(h)[:10])
                        if d < cutoff:
                            continue
                        stocks = [
                            str(row[ci]).strip()
                            for row in ws.iter_rows(min_row=2, values_only=True)
                            if ci < len(row) and row[ci]
                        ]
                        history[sheet_name][str(d)] = stocks
                    except Exception:
                        pass
        except Exception:
            pass

    # ── Merge today's results ──────────────────────────────────────────────────
    for sheet_name, df in results.items():
        if df is None or df.empty:
            continue
        scol = SCORE_COL.get(sheet_name, "Score /8")
        sorted_df = df.sort_values(scol, ascending=False)
        history[sheet_name][today_str] = [
            str(r.get("Stock", "")).strip()
            for _, r in sorted_df.iterrows()
            if str(r.get("Stock", "")).strip()
        ]

    # ── Build workbook ─────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F497D")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EEF3FB")
    CENTER    = Alignment(horizontal="center", vertical="center")
    THIN      = Border(
        left   = Side(style="thin", color="D0D0D0"),
        right  = Side(style="thin", color="D0D0D0"),
        top    = Side(style="thin", color="D0D0D0"),
        bottom = Side(style="thin", color="D0D0D0"),
    )
    TAB_COLORS = {
        "Probable Upside":        "4472C4",
        "Support Entry":          "70AD47",
        "Consolidation Breakout": "ED7D31",
    }

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, date_data in history.items():
        if not date_data:
            continue

        all_dates  = sorted(date_data.keys())   # oldest → newest
        max_stocks = max(len(v) for v in date_data.values())

        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = TAB_COLORS.get(sheet_name, "4472C4")
        ws.freeze_panes = "A2"

        # Row 1: date headers
        for ci, date_str in enumerate(all_dates, 1):
            cell = ws.cell(row=1, column=ci, value=date_str)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = CENTER
            cell.border    = THIN
        ws.row_dimensions[1].height = 22

        # Rows 2+: stock names ranked top→bottom per date column
        for ri in range(max_stocks):
            for ci, date_str in enumerate(all_dates, 1):
                stocks = date_data.get(date_str, [])
                val    = stocks[ri] if ri < len(stocks) else ""
                c = ws.cell(row=ri + 2, column=ci, value=val)
                c.alignment = CENTER
                c.border    = THIN
                c.font      = Font(bold=bool(val), size=10)
                if val and (ri + 2) % 2 == 0:
                    c.fill = ALT_FILL

        for ci in range(1, len(all_dates) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 14

    wb.save(HIST_FILE)

# ── Main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    log("NSE F&O Daily Export starting…")

    fno_live, fno_cached_names, fno_source = fetch_fno_symbols()
    COMPANY_NAMES_MERGED = {**fno_cached_names, **COMPANY_NAMES}
    FNO = (
        [(s, s, COMPANY_NAMES_MERGED.get(s, s)) for s in fno_live]
        if fno_live
        else [(s, s, n) for s, n in sorted(COMPANY_NAMES.items())]
    )
    log(f"F&O list: {len(FNO)} stocks ({fno_source})")

    tickers = [f"{s}.NS" for _, s, _ in FNO]

    log("Downloading 1Y price history for all stocks…")
    hist = yf.download(tickers, period="1y", auto_adjust=True, progress=False)
    close_df = hist["Close"]
    high_df  = hist["High"]
    low_df   = hist["Low"]
    vol_df   = hist["Volume"]
    log("Download complete. Running screeners…")

    df1 = run_probable_upside(FNO, close_df, vol_df, low_df)
    log(f"Probable Upside:        {len(df1)} candidates")

    df2 = run_support_entry(FNO, close_df, low_df, high_df, vol_df)
    log(f"Support Entry:          {len(df2)} candidates")

    df3 = run_consolidation(FNO, close_df, high_df, vol_df)
    log(f"Consolidation Breakout: {len(df3)} candidates")

    write_excel({
        "Probable Upside":        df1,
        "Support Entry":          df2,
        "Consolidation Breakout": df3,
    })
    log(f"Saved → {OUTFILE}")

    update_consolidated_history({
        "Probable Upside":        df1,
        "Support Entry":          df2,
        "Consolidation Breakout": df3,
    })
    log(f"History updated → {HIST_FILE}")

    # Cleanup: delete Excel files older than 20 days
    cutoff = datetime.date.today() - datetime.timedelta(days=20)
    deleted = []
    for fname in os.listdir(EXPORTS_DIR):
        if not fname.startswith("screener_") or not fname.endswith(".xlsx"):
            continue
        try:
            file_date = datetime.date.fromisoformat(fname[9:19])  # screener_YYYY-MM-DD.xlsx
            if file_date < cutoff:
                os.remove(os.path.join(EXPORTS_DIR, fname))
                deleted.append(fname)
        except Exception:
            pass
    if deleted:
        log(f"Deleted {len(deleted)} old file(s): {', '.join(deleted)}")

    # Cross-platform notification
    try:
        import platform, subprocess
        msg = (f"Screener export complete — "
               f"{len(df1)} upside, {len(df2)} support, {len(df3)} consolidation setups.")
        system = platform.system()
        if system == "Darwin":
            subprocess.run([
                "osascript", "-e",
                f'display notification "{msg}" with title "NSE F&O Export" sound name "Glass"'
            ], check=False)
        elif system == "Windows":
            subprocess.run([
                "powershell", "-Command",
                f'[System.Reflection.Assembly]::LoadWithPartialName("System.Windows.Forms") | Out-Null;'
                f'[System.Windows.Forms.MessageBox]::Show("{msg}", "NSE F&O Export")'
            ], check=False)
        # Linux: notification not needed (headless server)
    except Exception:
        pass
