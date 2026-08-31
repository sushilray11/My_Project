"""
Headless daily screener for Project Up.
Runs the same 5-step PVA pipeline as app.py — no Streamlit required.
Scheduled via cron at 15:00 Mon-Fri; output goes to daily_run.log.
"""
import sys, os, json, datetime
import pandas as pd
import requests, io

DIR = os.path.dirname(os.path.abspath(__file__))
UNIVERSE_CACHE = os.path.join(DIR, "universe_cache.json")
MCAP_CACHE     = os.path.join(DIR, "mcap_cache.json")
HIST_PATH      = os.path.join(DIR, "history.xlsx")

def _log(msg):
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def _load_json(path):
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_json(path, data):
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass

def _cache_age_days(data):
    try:
        return (datetime.date.today() - datetime.date.fromisoformat(data.get("updated", "2000-01-01"))).days
    except Exception:
        return 999

def _ema_series(prices, period):
    if len(prices) < period:
        return []
    k = 2.0 / (period + 1)
    val = sum(prices[:period]) / period
    out = [val]
    for p in prices[period:]:
        val = val * (1 - k) + p * k
        out.append(val)
    return out

def _ema_val(prices, period):
    s = _ema_series(prices, period)
    return s[-1] if s else None

def _fetch_universe():
    cache = _load_json(UNIVERSE_CACHE)
    if cache.get("symbols") and _cache_age_days(cache) < 7:
        _log(f"Universe: {len(cache['symbols'])} stocks (cache)")
        return cache["symbols"], cache.get("names", {})
    try:
        r = requests.get(
            "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
            timeout=25,
        )
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        sym_col  = df.columns[0]
        name_col = next((c for c in df.columns if "NAME" in c.upper()), None)
        symbols  = sorted(str(s).strip() for s in df[sym_col].dropna() if str(s).strip())
        names    = {}
        if name_col:
            for _, row in df.iterrows():
                s = str(row[sym_col]).strip()
                n = str(row[name_col]).strip()
                if s and n and n.lower() != "nan":
                    names[s] = n.title()
        if len(symbols) >= 100:
            _save_json(UNIVERSE_CACHE, {
                "updated": str(datetime.date.today()),
                "symbols": symbols,
                "names": names,
            })
            _log(f"Universe: {len(symbols)} stocks (live NSE)")
            return symbols, names
    except Exception as e:
        _log(f"Universe fetch error: {e}")
    if cache.get("symbols"):
        _log(f"Universe: {len(cache['symbols'])} stocks (offline cache)")
        return cache["symbols"], cache.get("names", {})
    return [], {}

def _get_mcap(sym_ns, mcap_data):
    today = str(datetime.date.today())
    entry = mcap_data.get(sym_ns)
    if entry:
        try:
            age = (datetime.date.today() - datetime.date.fromisoformat(entry["date"])).days
            if age < 7:
                return entry["mcap"]
        except Exception:
            pass
    try:
        import yfinance as yf
        mc = yf.Ticker(sym_ns).fast_info.market_cap or 0
    except Exception:
        mc = 0
    mcap_data[sym_ns] = {"mcap": mc, "date": today}
    return mc

def _write_formatted_excel(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HDR_FILL = PatternFill("solid", fgColor="1D4ED8")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="EFF6FF")
    SCR_HI   = PatternFill("solid", fgColor="D1FAE5")
    SCR_MED  = PatternFill("solid", fgColor="FEF9C3")
    THIN     = Side(style="thin", color="E2E8F0")
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER   = Alignment(horizontal="center", vertical="center")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df.empty:
            ws["A1"] = "No data"; continue
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                val = row[col]
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = BORDER; c.alignment = CENTER
                if ri % 2 == 0:
                    c.fill = ALT_FILL
                if col in ("Score /10", "Score"):
                    try:
                        sv = float(val)
                        c.fill = SCR_HI if sv >= 8 else SCR_MED
                        c.font = Font(bold=True, size=10)
                    except (TypeError, ValueError):
                        pass
        for ci, col in enumerate(cols, 1):
            ml = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 28)
    wb.save(path)

def _save_history(rows, top_n=20):
    if not rows:
        return
    # use last market-open day so weekend runs don't create non-trading dates
    d = datetime.date.today()
    while d.weekday() >= 5:   # 5=Sat, 6=Sun
        d -= datetime.timedelta(days=1)
    today = str(d)
    df_new = (
        pd.DataFrame(rows)
        .sort_values(["Score /10", "Vol Ratio", "Mkt Cap (Cr)"], ascending=False)
        .head(top_n)
    )
    df_new.insert(0, "Date", today)
    if os.path.exists(HIST_PATH):
        try:
            existing = pd.read_excel(HIST_PATH, dtype=str)
            existing = existing[existing["Date"].astype(str) != today]
            combined = pd.concat([existing, df_new.astype(str)], ignore_index=True)
        except Exception:
            combined = df_new.astype(str)
    else:
        combined = df_new.astype(str)
    _write_formatted_excel(HIST_PATH, {"History": combined})
    _log(f"Saved {len(df_new)} rows to history.xlsx")

def run():
    import yfinance as yf

    _log("Project Up — daily screener starting")

    # ── Step 1: Universe ──────────────────────────────────────────────────────
    symbols, names = _fetch_universe()
    if not symbols:
        _log("ERROR: empty universe — aborting")
        sys.exit(1)

    tickers = [f"{s}.NS" for s in symbols]

    # ── Step 2: Batch OHLCVO download ─────────────────────────────────────────
    BATCH = 400
    all_close, all_high, all_low, all_vol, all_open = {}, {}, {}, {}, {}
    batches = [tickers[i:i+BATCH] for i in range(0, len(tickers), BATCH)]
    _log(f"Downloading {len(tickers)} tickers in {len(batches)} batches…")

    for bi, batch in enumerate(batches):
        _log(f"  Batch {bi+1}/{len(batches)} ({len(batch)} stocks)")
        try:
            dh = yf.download(batch, period="1y", auto_adjust=True, progress=False, threads=True)
            if dh.empty:
                continue
            for field, store in [("Close", all_close), ("High", all_high),
                                  ("Low", all_low), ("Volume", all_vol), ("Open", all_open)]:
                try:
                    sub = dh[field]
                except KeyError:
                    continue
                if isinstance(sub, pd.Series):
                    s = sub.dropna()
                    if len(s) >= 60:
                        store[batch[0]] = list(s.astype(float))
                else:
                    for col in sub.columns:
                        s = sub[col].dropna()
                        if len(s) >= 60:
                            store[col] = list(s.astype(float))
        except Exception as e:
            _log(f"  Batch {bi+1} error: {e}")

    # ── Step 3: 200 EMA filter ────────────────────────────────────────────────
    ema200_pass = [
        sym_ns for sym_ns, c in all_close.items()
        if len(c) >= 200 and _ema_val(c, 200) and c[-1] > _ema_val(c, 200)
    ]
    _log(f"200 EMA filter: {len(ema200_pass)} pass")

    # ── Step 4: Market cap filter ─────────────────────────────────────────────
    mcap_data = _load_json(MCAP_CACHE)
    mcap_pass = []
    for mi, sym_ns in enumerate(ema200_pass):
        mc = _get_mcap(sym_ns, mcap_data)
        if mc >= 10_000_000_000:
            mcap_pass.append((sym_ns, mc))
        if (mi + 1) % 100 == 0:
            _save_json(MCAP_CACHE, mcap_data)
            _log(f"  Mcap check {mi+1}/{len(ema200_pass)}: {len(mcap_pass)} qualifying")
    _save_json(MCAP_CACHE, mcap_data)
    _log(f"Mkt Cap filter: {len(mcap_pass)} pass")

    # ── Step 5: Score setup signals ───────────────────────────────────────────
    rows = []
    for sym_ns, mc in mcap_pass:
        try:
            sym = sym_ns.replace(".NS", "")
            c = all_close[sym_ns]
            h = all_high.get(sym_ns, [])
            l = all_low.get(sym_ns, [])
            v = all_vol.get(sym_ns, [])

            if len(c) < 60 or len(h) < 11 or len(l) < 11 or len(v) < 20:
                continue

            price = c[-1]
            e200  = _ema_val(c, 200)
            e20   = _ema_val(c, 20)
            e50   = _ema_val(c, 50)
            if not (e200 and e20 and e50):
                continue
            above_200_pct = round((price - e200) / e200 * 100, 1)

            avgv20    = sum(v[-20:]) / 20
            vol_ratio = round(v[-1] / avgv20, 2) if avgv20 else 0.0
            day_range = max(h[-1] - l[-1], 0.01)

            vol_dryup  = bool(avgv20 and v[-1] < avgv20 * 0.70)
            vol_low5   = bool(len(v) >= 5 and v[-1] <= min(v[-5:]))
            squeeze3   = bool(len(h) >= 3 and len(l) >= 3 and
                              (max(h[-3:]) - min(l[-3:])) / price < 0.04)
            prior_up   = bool(len(c) >= 16 and c[-1] > c[-16])

            up_v10 = dn_v10 = 0.0
            for i in range(max(1, len(c) - 10), len(c)):
                if c[i] > c[i-1]:
                    up_v10 += v[i]
                else:
                    dn_v10 += v[i]
            net_accum = up_v10 > dn_v10 * 1.2

            weak_selling = True
            for i in range(max(1, len(c) - 5), len(c)):
                if c[i] < c[i-1] and avgv20 and v[i] >= avgv20 * 0.80:
                    weak_selling = False
                    break

            rising_lows    = bool(len(l) >= 7 and l[-1] > l[-4] and l[-4] > l[-7])
            avg_close_pos  = sum((c[i]-l[i])/max(h[i]-l[i], 0.01) for i in range(-3, 0)) / 3
            bullish_closes = avg_close_pos >= 0.55
            at_support     = ((-0.02 <= (price/e20 - 1) <= 0.04) or
                              (-0.01 <= (price/e50 - 1) <= 0.03))
            entry_trigger  = bool(len(v) >= 2 and len(c) >= 2 and
                                  v[-1] > v[-2] and c[-1] > c[-2])

            score = sum([vol_dryup, vol_low5, squeeze3, prior_up,
                         net_accum, weak_selling, rising_lows,
                         bullish_closes, at_support, entry_trigger])

            if score >= 5:
                high52 = max(h[-252:]) if len(h) >= 252 else max(h)
                low5   = min(l[-5:]) if len(l) >= 5 else l[-1]
                buy_lo = round(price * 0.995, 2)
                buy_hi = round(price * 1.005, 2)
                stop   = round(low5 * 0.99, 2)
                target = round(min(price * 1.07, high52 * 0.995), 2)
                risk   = round((price - stop) / price * 100, 1)
                reward = round((target - price) / price * 100, 1)
                rr     = round(reward / risk, 1) if risk > 0 else 0.0

                rows.append({
                    "Stock":          sym,
                    "Company":        names.get(sym, sym),
                    "Price (₹)":      round(price, 2),
                    "Mkt Cap (Cr)":   round(mc / 1e7),
                    "Above 200 EMA%": above_200_pct,
                    "Vol Ratio":      vol_ratio,
                    "Prior Uptrend":  "Yes" if prior_up      else "No",
                    "3D Squeeze":     "Yes" if squeeze3      else "No",
                    "Net Accum":      "Yes" if net_accum     else "No",
                    "Rising Lows":    "Yes" if rising_lows   else "No",
                    "Entry Trigger":  "Yes" if entry_trigger else "No",
                    "Score /10":      score,
                    "Buy Zone":       f"{buy_lo}-{buy_hi}",
                    "Target (₹)":    target,
                    "Stop (₹)":      stop,
                    "R:R":            rr,
                    "Chart":          f"https://www.tradingview.com/chart/?symbol=NSE:{sym}",
                })
        except Exception:
            pass

    _log(f"Scoring complete: {len(rows)} stocks qualify (score ≥ 5)")
    _save_history(rows)

    top = sorted(rows, key=lambda r: (r["Score /10"], r["Vol Ratio"]), reverse=True)[:20]
    _log("Top 20 picks:")
    for i, r in enumerate(top, 1):
        _log(f"  {i:2d}. {r['Stock']:12s}  Score {r['Score /10']}/10  Vol {r['Vol Ratio']:.1f}x  Price ₹{r['Price (₹)']}")

if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        _log(f"FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
