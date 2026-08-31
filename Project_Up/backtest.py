"""
Standalone backtest for Project Up picks.
Reads history.xlsx, takes last 10 pick-dates, downloads 3-month price data,
and calculates D+1 / D+3 / D+5 forward returns for each pick.
Saves backtest_results.xlsx with two sheets: Pick Results and Summary.
"""
import os, sys, datetime
import pandas as pd

DIR       = os.path.dirname(os.path.abspath(__file__))
HIST_PATH = os.path.join(DIR, "history.xlsx")
OUT_PATH  = os.path.join(DIR, "backtest_results.xlsx")

def _write_formatted_excel(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HDR_FILL = PatternFill("solid", fgColor="1D4ED8")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL = PatternFill("solid", fgColor="EFF6FF")
    POS_FILL = PatternFill("solid", fgColor="DCFCE7")
    NEG_FILL = PatternFill("solid", fgColor="FEE2E2")
    SCR_HI   = PatternFill("solid", fgColor="D1FAE5")
    SCR_MED  = PatternFill("solid", fgColor="FEF9C3")
    THIN     = Side(style="thin", color="E2E8F0")
    BORDER   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER   = Alignment(horizontal="center", vertical="center")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df.empty:
            ws["A1"] = "No data"; continue
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                val = row[col]
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = BORDER; c.alignment = CENTER
                if ri % 2 == 0:
                    c.fill = ALT_FILL
                if "%" in str(col) and col not in ("Score /10",):
                    try:
                        fv = float(val)
                        c.fill = POS_FILL if fv > 0 else NEG_FILL
                    except (TypeError, ValueError):
                        pass
                if col in ("Score /10", "Score"):
                    try:
                        sv = float(val)
                        c.fill = SCR_HI if sv >= 8 else SCR_MED
                        c.font = Font(bold=True, size=10)
                    except (TypeError, ValueError):
                        pass
        for ci in range(1, len(cols) + 1):
            ml = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 28)
    wb.save(path)

def _log(msg):
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def _hr(series):
    v = series.dropna()
    return f"{(v > 0).mean()*100:.1f}%" if len(v) else "N/A"

def _avg(series):
    v = series.dropna()
    return f"{v.mean():.2f}%" if len(v) else "N/A"

def _stats(df, label):
    r = {"Group": label, "Picks": len(df)}
    for h, col in [("1D", "D+1 %"), ("3D", "D+3 %"), ("5D", "D+5 %")]:
        v = df[col].dropna()
        r[f"Hit Rate {h}"]   = f"{(v > 0).mean()*100:.1f}%" if len(v) else "N/A"
        r[f"Avg Return {h}"] = f"{v.mean():.2f}%"           if len(v) else "N/A"
    return r

def run():
    import yfinance as yf

    _log("Project Up — Backtest starting")

    if not os.path.exists(HIST_PATH):
        _log("ERROR: history.xlsx not found. Run the screener first.")
        sys.exit(1)

    raw = pd.read_excel(HIST_PATH, dtype=str)
    raw["Date"]      = pd.to_datetime(raw["Date"], errors="coerce")
    raw["Score /10"] = pd.to_numeric(raw["Score /10"], errors="coerce")
    raw["Price (₹)"] = pd.to_numeric(raw["Price (₹)"], errors="coerce")
    raw = raw.dropna(subset=["Date", "Stock"])

    dates_avail = sorted(raw["Date"].unique())[-10:]
    raw = raw[raw["Date"].isin(dates_avail)]
    _log(f"Using {len(dates_avail)} pick-dates, {len(raw)} total picks")

    bt_tickers = [f"{s}.NS" for s in raw["Stock"].unique().tolist()]
    _log(f"Downloading 3-month price data for {len(bt_tickers)} stocks…")

    try:
        dh = yf.download(bt_tickers, period="3mo", auto_adjust=True,
                         progress=False, threads=True)
        bt_close = {}
        try:
            sub = dh["Close"]
        except KeyError:
            sub = None
        if sub is not None:
            if isinstance(sub, pd.Series):
                bt_close[bt_tickers[0]] = sub.dropna()
            else:
                for col in sub.columns:
                    s = sub[col].dropna()
                    if len(s) > 0:
                        bt_close[col] = s
    except Exception as e:
        _log(f"Download error: {e}")
        bt_close = {}

    today = pd.Timestamp.today().normalize()
    bt_rows = []

    for _, row in raw.iterrows():
        pick_date = row["Date"]
        sym       = row["Stock"]
        tkr       = f"{sym}.NS"

        if tkr not in bt_close:
            continue

        prices     = bt_close[tkr].sort_index()
        dates_list = list(prices.index)
        # last trading day on or before pick_date (handles weekend picks)
        pick_pos = None
        for i, d in enumerate(dates_list):
            if d <= pick_date:
                pick_pos = i
        if pick_pos is None:
            continue
        pick_px = float(prices.iloc[pick_pos])

        def _fwd(offset):
            idx = pick_pos + offset
            if idx >= len(dates_list) or dates_list[idx] > today:
                return None, None
            p = float(prices.iloc[idx])
            return round(p, 2), round((p - pick_px) / pick_px * 100, 2)

        p1, r1 = _fwd(1)
        p3, r3 = _fwd(3)
        p5, r5 = _fwd(5)

        bt_rows.append({
            "Date":    str(pick_date.date()),
            "Stock":   sym,
            "Company": row.get("Company", sym),
            "Score":   row.get("Score /10"),
            "Pick ₹":  round(pick_px, 2),
            "D+1 ₹":   p1,  "D+1 %": r1,
            "D+3 ₹":   p3,  "D+3 %": r3,
            "D+5 ₹":   p5,  "D+5 %": r5,
        })

    if not bt_rows:
        _log("No results — not enough forward data yet (need at least D+1 after pick date).")
        sys.exit(0)

    bt_df = pd.DataFrame(bt_rows)
    bt_df["Score"] = pd.to_numeric(bt_df["Score"], errors="coerce")

    _log(f"\n{'='*60}")
    _log(f"BACKTEST SUMMARY  ({len(bt_df)} picks across {bt_df['Date'].nunique()} dates)")
    _log(f"{'='*60}")
    _log(f"Hit Rate  — 1D: {_hr(bt_df['D+1 %'])}  3D: {_hr(bt_df['D+3 %'])}  5D: {_hr(bt_df['D+5 %'])}")
    _log(f"Avg Return— 1D: {_avg(bt_df['D+1 %'])}  3D: {_avg(bt_df['D+3 %'])}  5D: {_avg(bt_df['D+5 %'])}")

    hi = bt_df[bt_df["Score"] >= 8]
    lo = bt_df[bt_df["Score"] <  8]
    if len(hi):
        _log(f"Score ≥ 8 ({len(hi)} picks): 1D {_hr(hi['D+1 %'])}  3D {_hr(hi['D+3 %'])}  5D {_hr(hi['D+5 %'])}")
    if len(lo):
        _log(f"Score 5-7 ({len(lo)} picks): 1D {_hr(lo['D+1 %'])}  3D {_hr(lo['D+3 %'])}  5D {_hr(lo['D+5 %'])}")
    _log(f"{'='*60}\n")

    RET_COLS = ["D+1 ₹","D+3 ₹","D+5 ₹","D+1 %","D+3 %","D+5 %"]
    KEY_COLS = ["Date", "Stock"]

    def _upsert_up(existing, new):
        if existing is None or len(existing) == 0:
            return new.copy()
        for df in [existing, new]:
            for c in KEY_COLS:
                if c in df.columns:
                    df[c] = df[c].astype(str).str.strip()
        key_idx = {(r["Date"], r["Stock"]): i for i, r in existing.iterrows()}
        new_rows = []
        for _, row in new.iterrows():
            k = (str(row["Date"]), str(row["Stock"]))
            if k in key_idx:
                ei = key_idx[k]
                for col in RET_COLS:
                    if col in row and col in existing.columns:
                        v = row[col]
                        if v is not None and str(v) not in ("—", "nan", "None", ""):
                            existing.at[ei, col] = v
            else:
                new_rows.append(row.to_dict())
        if new_rows:
            existing = pd.concat([existing, pd.DataFrame(new_rows)], ignore_index=True)
        return existing

    if os.path.exists(OUT_PATH):
        try:
            _existing = pd.read_excel(OUT_PATH, sheet_name="Pick Results", dtype=str)
            combined_bt = _upsert_up(_existing, bt_df.astype(str))
        except Exception:
            combined_bt = bt_df.astype(str)
    else:
        combined_bt = bt_df.astype(str)

    for col in ["D+1 %", "D+3 %", "D+5 %"]:
        combined_bt[col] = pd.to_numeric(combined_bt[col], errors="coerce")

    # keep only last 30 days
    combined_bt["Date"] = pd.to_datetime(combined_bt["Date"], errors="coerce")
    cutoff = pd.Timestamp.today() - pd.Timedelta(days=30)
    combined_bt = combined_bt[combined_bt["Date"] >= cutoff]
    combined_bt["Date"] = combined_bt["Date"].dt.strftime("%Y-%m-%d")

    hi_c = combined_bt[pd.to_numeric(combined_bt.get("Score", pd.Series()), errors="coerce") >= 8]
    lo_c = combined_bt[pd.to_numeric(combined_bt.get("Score", pd.Series()), errors="coerce") <  8]

    summary_rows = [_stats(bt_df, "All Picks")]
    if len(hi_c): summary_rows.append(_stats(hi_c, "Score ≥ 8"))
    if len(lo_c): summary_rows.append(_stats(lo_c, "Score 5–7"))

    _write_formatted_excel(OUT_PATH, {
        "Pick Results": combined_bt,
        "Summary":      pd.DataFrame(summary_rows),
    })

    _log(f"Saved → {OUT_PATH}  ({len(combined_bt)} total rows)")

if __name__ == "__main__":
    try:
        run()
    except Exception as e:
        _log(f"FATAL ERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
