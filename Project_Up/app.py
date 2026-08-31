import streamlit as st
import pandas as pd
import json as _json
import os as _os
import datetime as _dt

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Project Up — Short-Term Stock Screener",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ─────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

html, body, [data-testid="stAppViewContainer"], .main {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif !important;
}
.main { background: #f1f5f9 !important; }
.main .block-container {
    padding-top: 0 !important; padding-left: 0 !important;
    padding-right: 0 !important; padding-bottom: 0 !important;
    max-width: 100% !important;
}
[data-testid="stHeader"]    { display: none !important; height: 0 !important; }
[data-testid="stToolbar"]   { display: none !important; }
[data-testid="stDecoration"]{ display: none !important; }
header { display: none !important; height: 0 !important; }
:root  { --header-height: 0rem !important; }
section.main > div { padding-top: 0 !important; margin-top: 0 !important; }
.appview-container .main .block-container { padding-top: 0 !important; }
[data-testid="stMainBlockContainer"] { padding-top: 0 !important; }
div.block-container { padding-top: 0 !important; margin-top: 0 !important; }
.stApp > div:first-child { padding-top: 0 !important; margin-top: 0 !important; }
.top-navbar { margin-top: -2rem !important; }
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stVerticalBlock"] > div { gap: 0 !important; }

[data-testid="stSidebar"] {
    background: #0f172a !important;
    border-right: 1px solid #1e293b !important;
}
[data-testid="stSidebar"] p, [data-testid="stSidebar"] label,
[data-testid="stSidebar"] span, [data-testid="stSidebar"] small,
[data-testid="stSidebar"] .stCaption { color: #94a3b8 !important; }
[data-testid="stSidebar"] h1, [data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3 { color: #f1f5f9 !important; }

.top-navbar {
    background: linear-gradient(135deg, #052e16 0%, #064e3b 100%);
    padding: 1.1rem 2.2rem;
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 0;
}
.nav-title   { font-size: 1.45rem; font-weight: 700; color: #f8fafc; letter-spacing: -0.02em; margin: 0; }
.nav-subtitle{ font-size: 0.78rem; color: #94a3b8; margin-top: 3px; }
.nav-badge   {
    background: #10b981; color: #ecfdf5;
    padding: 5px 14px; border-radius: 20px;
    font-size: 0.75rem; font-weight: 600; letter-spacing: 0.03em;
}
.content-wrap { padding: 0.6rem 2rem; }

div[data-testid="metric-container"] {
    background: white; border-radius: 8px;
    padding: 0.4rem 0.8rem !important;
    box-shadow: 0 1px 3px rgba(0,0,0,0.08);
    border-left: 3px solid #10b981;
}
div[data-testid="metric-container"] label,
div[data-testid="metric-container"] [data-testid="stMetricLabel"] p {
    color: #64748b !important; font-size: 0.68rem !important;
    text-transform: uppercase; letter-spacing: 0.05em; font-weight: 500 !important;
}
div[data-testid="metric-container"] [data-testid="stMetricValue"],
div[data-testid="metric-container"] [data-testid="stMetricValue"] > div {
    color: #0f172a !important; font-size: 0.9rem !important; font-weight: 700 !important; line-height: 1.3 !important;
}

.section-header {
    background: #0f172a; padding: 0.85rem 1.4rem;
    border-radius: 10px 10px 0 0;
    display: flex; align-items: center; justify-content: space-between;
    margin-top: 0.5rem;
}
.section-header-title { color: #f8fafc; font-size: 0.95rem; font-weight: 600; }
.section-badge {
    background: #10b981; color: #ecfdf5;
    padding: 3px 10px; border-radius: 12px; font-size: 0.72rem; font-weight: 600;
}
.section-body {
    background: white; padding: 1rem 1.2rem;
    border-radius: 0 0 10px 10px;
    border: 1px solid #e2e8f0; border-top: none;
    margin-bottom: 1.5rem;
}

div[data-testid="stDataFrame"] { border-radius: 8px; overflow: hidden; }
div.stButton > button {
    background: linear-gradient(135deg, #059669, #10b981) !important;
    color: white !important; border: none !important;
    padding: 0.5rem 1.5rem !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 0.88rem !important;
    transition: all 0.2s ease !important;
}
div.stButton > button:hover { opacity: 0.9 !important; transform: translateY(-1px) !important; }
div[data-testid="stProgressBar"] > div { background: #10b981 !important; }
hr { border-color: #e2e8f0 !important; margin: 1.2rem 0 !important; }
</style>
""", unsafe_allow_html=True)

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:1.2rem 0.5rem 1rem; border-bottom:1px solid #1e293b; margin-bottom:1rem;">
        <div style="font-size:1.1rem;font-weight:700;color:#f8fafc;letter-spacing:-0.02em;">🚀 Project Up</div>
        <div style="font-size:0.72rem;color:#64748b;margin-top:3px;">Short-Term Bullish Screener</div>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("""
    <div style="padding:0.5rem;font-size:0.78rem;color:#94a3b8;line-height:1.6;">
    <b style="color:#f1f5f9;">How it works</b><br><br>
    ① Fetches all NSE equity stocks from NSE<br>
    ② Filters: Market Cap > ₹1000 Cr + Price > 200 EMA<br>
    ③ Scores 10 Price-Volume Action (PVA) signals<br>
    ④ Shows top 20 with highest probability to go up in 1–7 days
    <br><br>
    <b style="color:#f1f5f9;">PVA signals:</b> candle patterns, volume confirmation, accumulation — no lagging indicators (RSI/MACD/EMA).
    <br><br>
    <b style="color:#f1f5f9;">First run</b> downloads market cap for ~500–700 stocks (~3–5 min). Subsequent runs use the 7-day cache.
    </div>
    """, unsafe_allow_html=True)

# ── Top Navbar ─────────────────────────────────────────────────────────────────
st.markdown("""
<div class="top-navbar">
    <div>
        <div class="nav-title">🚀 Project Up</div>
        <div class="nav-subtitle">NSE Short-Term Bullish Screener — <span style="font-size:0.72rem;opacity:0.7;">powered by Zerodha + yfinance</span></div>
    </div>
    <span class="nav-badge">Mkt Cap &gt; ₹1000 Cr · Above 200 EMA</span>
</div>
<div class="content-wrap" style="padding-bottom:0">
""", unsafe_allow_html=True)

# ── Cache helpers ──────────────────────────────────────────────────────────────
_DIR = _os.path.dirname(_os.path.abspath(__file__))
_UNIVERSE_CACHE = _os.path.join(_DIR, "universe_cache.json")
_MCAP_CACHE     = _os.path.join(_DIR, "mcap_cache.json")

def _load_json(path):
    try:
        with open(path) as f:
            return _json.load(f)
    except Exception:
        return {}

def _save_json(path, data):
    try:
        with open(path, "w") as f:
            _json.dump(data, f)
    except Exception:
        pass

def _cache_age_days(data):
    try:
        return (_dt.date.today() - _dt.date.fromisoformat(data.get("updated", "2000-01-01"))).days
    except Exception:
        return 999

# ── Universe fetching ──────────────────────────────────────────────────────────
@st.cache_data(ttl=86400, show_spinner=False)
def _fetch_universe():
    cache = _load_json(_UNIVERSE_CACHE)
    if cache.get("symbols") and _cache_age_days(cache) < 7:
        return cache["symbols"], cache.get("names", {}), "Cache"

    try:
        import requests, io
        # NSE equity master — clean list of all NSE-listed equity stocks (no bonds/NCDs)
        r = requests.get(
            "https://archives.nseindia.com/content/equities/EQUITY_L.csv",
            timeout=25,
        )
        r.raise_for_status()
        df = pd.read_csv(io.StringIO(r.text))
        sym_col  = df.columns[0]
        name_col = next((c for c in df.columns if "NAME" in c.upper()), None)
        symbols = sorted(str(s).strip() for s in df[sym_col].dropna() if str(s).strip())
        names = {}
        if name_col:
            for _, row in df.iterrows():
                s = str(row[sym_col]).strip()
                n = str(row[name_col]).strip()
                if s and n and n.lower() != "nan":
                    names[s] = n.title()
        if len(symbols) >= 100:
            _save_json(_UNIVERSE_CACHE, {
                "updated": str(_dt.date.today()),
                "symbols": symbols,
                "names": names,
            })
            return symbols, names, "NSE Equity List (Live)"
    except Exception:
        pass

    if cache.get("symbols"):
        return cache["symbols"], cache.get("names", {}), "Cache (offline)"
    return [], {}, "Error"

# ── Technical indicator helpers ────────────────────────────────────────────────
def _ema_series(prices, period):
    if len(prices) < period:
        return []
    k = 2.0 / (period + 1)
    val = sum(prices[:period]) / period
    out = [val]
    for p in prices[period:]:
        val = val * (1 - k) + p * k
        out.append(val)
    return out

def _ema_val(prices, period):
    s = _ema_series(prices, period)
    return s[-1] if s else None

def _rsi_wilder(prices, period=14):
    if len(prices) < period + 1:
        return 50.0
    gains  = [max(prices[i] - prices[i-1], 0) for i in range(1, len(prices))]
    losses = [max(prices[i-1] - prices[i], 0) for i in range(1, len(prices))]
    ag = sum(gains[:period])  / period
    al = sum(losses[:period]) / period
    for g, l in zip(gains[period:], losses[period:]):
        ag = (ag * (period - 1) + g) / period
        al = (al * (period - 1) + l) / period
    return 100 - (100 / (1 + ag / max(al, 1e-10)))

def _macd(prices):
    """Returns (macd_val, signal_val, hist_now, hist_prev)."""
    e12 = _ema_series(prices, 12)
    e26 = _ema_series(prices, 26)
    if len(e12) < 1 or len(e26) < 1:
        return None, None, 0, 0
    n = min(len(e12), len(e26))
    macd_line = [a - b for a, b in zip(e12[-n:], e26[-n:])]
    sig = _ema_series(macd_line, 9)
    if not sig:
        return macd_line[-1], None, 0, 0
    hist = [m - s for m, s in zip(macd_line[-len(sig):], sig)]
    h_now  = hist[-1]  if hist       else 0
    h_prev = hist[-2]  if len(hist) >= 2 else 0
    return macd_line[-1], sig[-1], h_now, h_prev

# ── Market cap helpers ─────────────────────────────────────────────────────────
def _get_mcap(sym_ns, mcap_data):
    today = str(_dt.date.today())
    entry = mcap_data.get(sym_ns)
    if entry:
        try:
            age = (_dt.date.today() - _dt.date.fromisoformat(entry["date"])).days
            if age < 7:
                return entry["mcap"]
        except Exception:
            pass
    try:
        import yfinance as yf
        fi = yf.Ticker(sym_ns).fast_info
        mc = fi.market_cap or 0
    except Exception:
        mc = 0
    mcap_data[sym_ns] = {"mcap": mc, "date": today}
    return mc

# ── History export helper ──────────────────────────────────────────────────────
def _write_formatted_excel(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HDR_FILL  = PatternFill("solid", fgColor="1D4ED8")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    ALT_FILL  = PatternFill("solid", fgColor="EFF6FF")
    POS_FILL  = PatternFill("solid", fgColor="DCFCE7")
    NEG_FILL  = PatternFill("solid", fgColor="FEE2E2")
    SCR_HI    = PatternFill("solid", fgColor="D1FAE5")
    SCR_MED   = PatternFill("solid", fgColor="FEF9C3")
    THIN      = Side(style="thin", color="E2E8F0")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER    = Alignment(horizontal="center", vertical="center")
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if df.empty:
            ws["A1"] = "No data"; continue
        cols = list(df.columns)
        for ci, col in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.fill = HDR_FILL; c.font = HDR_FONT
            c.alignment = CENTER; c.border = BORDER
        ws.row_dimensions[1].height = 24
        ws.freeze_panes = "A2"
        for ri, (_, row) in enumerate(df.iterrows(), 2):
            for ci, col in enumerate(cols, 1):
                val = row[col]
                c = ws.cell(row=ri, column=ci, value=val)
                c.border = BORDER; c.alignment = CENTER
                if ri % 2 == 0:
                    c.fill = ALT_FILL
                if "%" in str(col) and col not in ("Score /10",):
                    try:
                        fv = float(val)
                        c.fill = POS_FILL if fv > 0 else NEG_FILL
                    except (TypeError, ValueError):
                        pass
                if col in ("Score /10", "Score"):
                    try:
                        sv = float(val)
                        c.fill = SCR_HI if sv >= 8 else SCR_MED
                        c.font = Font(bold=True, size=10)
                    except (TypeError, ValueError):
                        pass
        for ci, col in enumerate(cols, 1):
            ml = max((len(str(ws.cell(r, ci).value or "")) for r in range(1, ws.max_row + 1)), default=8)
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 28)
    wb.save(path)

def _save_history(rows, top_n=20):
    if not rows:
        return
    # use last market-open day so weekend runs don't create non-trading dates
    d = _dt.date.today()
    while d.weekday() >= 5:   # 5=Sat, 6=Sun
        d -= _dt.timedelta(days=1)
    today = str(d)
    hist_path = _os.path.join(_DIR, "history.xlsx")
    df_new = (
        pd.DataFrame(rows)
        .sort_values(["Score /10", "Vol Ratio", "Mkt Cap (Cr)"], ascending=False)
        .head(top_n)
    )
    df_new.insert(0, "Date", today)
    if _os.path.exists(hist_path):
        try:
            existing = pd.read_excel(hist_path, dtype=str)
            existing = existing[existing["Date"].astype(str) != today]
            combined = pd.concat([existing, df_new.astype(str)], ignore_index=True)
        except Exception:
            combined = df_new.astype(str)
    else:
        combined = df_new.astype(str)
    try:
        _write_formatted_excel(hist_path, {"History": combined})
    except Exception:
        combined.to_excel(hist_path, index=False)

# ── Main UI ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="section-header">
    <span class="section-header-title">🚀 Project Up — NSE Setup Screener</span>
    <span class="section-badge">Top 20 setups · Backtest included</span>
</div>
<div class="section-body">
""", unsafe_allow_html=True)

tab1, tab2 = st.tabs(["🚀 Screener", "📊 Backtest (Last 10 Days)"])

# ── TAB 1: SCREENER ────────────────────────────────────────────────────────────
with tab1:
    st.caption(
        "Screens all NSE stocks (Mkt Cap > ₹1000 Cr, Price > 200 EMA) and finds stocks "
        "consolidating quietly — low volume, tight range, sellers absent — before the next leg up. "
        "Best for 1–7 day upside trades."
    )

    if st.button("🚀 Find Top Upside Setups", key="run_screener"):
        st.session_state.pop("up_data", None)
        st.session_state["up_requested"] = True

    if st.session_state.get("up_requested"):
        if "up_data" not in st.session_state:
            import yfinance as yf

            with st.spinner("Fetching stock universe from NSE…"):
                symbols, names, src = _fetch_universe()

            if not symbols:
                st.error("Could not fetch stock universe. Check your internet connection.")
                st.stop()

            tickers = [f"{s}.NS" for s in symbols]
            st.info(f"Universe: **{len(tickers)} NSE equity stocks** ({src}). Downloading 1Y price history in batches…")

            BATCH = 400
            all_close, all_high, all_low, all_vol, all_open = {}, {}, {}, {}, {}
            prog_dl = st.progress(0, text="Downloading price data…")
            batches = [tickers[i:i+BATCH] for i in range(0, len(tickers), BATCH)]

            for bi, batch in enumerate(batches):
                prog_dl.progress((bi + 1) / len(batches), text=f"Batch {bi+1}/{len(batches)} ({len(batch)} stocks)…")
                try:
                    dh = yf.download(batch, period="1y", auto_adjust=True, progress=False, threads=True)
                    if dh.empty:
                        continue
                    for field, store in [("Close", all_close), ("High", all_high),
                                         ("Low", all_low), ("Volume", all_vol), ("Open", all_open)]:
                        try:
                            sub = dh[field]
                        except KeyError:
                            continue
                        if isinstance(sub, pd.Series):
                            s = sub.dropna()
                            if len(s) >= 60:
                                store[batch[0]] = list(s.astype(float))
                        else:
                            for col in sub.columns:
                                s = sub[col].dropna()
                                if len(s) >= 60:
                                    store[col] = list(s.astype(float))
                except Exception:
                    pass

            prog_dl.empty()

            ema200_pass = []
            for sym_ns in all_close:
                c = all_close[sym_ns]
                if len(c) < 200:
                    continue
                e200 = _ema_val(c, 200)
                if e200 and c[-1] > e200:
                    ema200_pass.append(sym_ns)

            st.info(f"**{len(ema200_pass)} stocks** pass the 200 EMA filter. Fetching market cap data (cached 7 days)…")

            mcap_data = _load_json(_MCAP_CACHE)
            mcap_pass = []
            prog_mc = st.progress(0, text="Checking market caps…")

            for mi, sym_ns in enumerate(ema200_pass):
                prog_mc.progress((mi + 1) / len(ema200_pass), text=f"Market cap {mi+1}/{len(ema200_pass)}: {sym_ns}…")
                mc = _get_mcap(sym_ns, mcap_data)
                if mc >= 10_000_000_000:
                    mcap_pass.append((sym_ns, mc))
                if (mi + 1) % 50 == 0:
                    _save_json(_MCAP_CACHE, mcap_data)

            _save_json(_MCAP_CACHE, mcap_data)
            prog_mc.empty()

            # ── Step 5: Score setup signals ───────────────────────────────────────
            rows = []
            for sym_ns, mc in mcap_pass:
                try:
                    sym = sym_ns.replace(".NS", "")
                    c = all_close[sym_ns]
                    h = all_high.get(sym_ns, [])
                    l = all_low.get(sym_ns, [])
                    v = all_vol.get(sym_ns, [])

                    if len(c) < 60 or len(h) < 11 or len(l) < 11 or len(v) < 20:
                        continue

                    price = c[-1]
                    e200  = _ema_val(c, 200)
                    e20   = _ema_val(c, 20)
                    e50   = _ema_val(c, 50)
                    if not (e200 and e20 and e50):
                        continue
                    above_200_pct = round((price - e200) / e200 * 100, 1)

                    avgv20    = sum(v[-20:]) / 20
                    vol_ratio = round(v[-1] / avgv20, 2) if avgv20 else 0.0
                    day_range = max(h[-1] - l[-1], 0.01)

                    vol_dryup  = bool(avgv20 and v[-1] < avgv20 * 0.70)
                    vol_low5   = bool(len(v) >= 5 and v[-1] <= min(v[-5:]))
                    squeeze3   = bool(len(h) >= 3 and len(l) >= 3 and
                                      (max(h[-3:]) - min(l[-3:])) / price < 0.04)
                    prior_up   = bool(len(c) >= 16 and c[-1] > c[-16])

                    up_v10 = dn_v10 = 0.0
                    for i in range(max(1, len(c) - 10), len(c)):
                        if c[i] > c[i-1]:
                            up_v10 += v[i]
                        else:
                            dn_v10 += v[i]
                    net_accum = up_v10 > dn_v10 * 1.2

                    weak_selling = True
                    for i in range(max(1, len(c) - 5), len(c)):
                        if c[i] < c[i-1] and avgv20 and v[i] >= avgv20 * 0.80:
                            weak_selling = False
                            break

                    rising_lows    = bool(len(l) >= 7 and l[-1] > l[-4] and l[-4] > l[-7])
                    avg_close_pos  = sum(
                        (c[i] - l[i]) / max(h[i] - l[i], 0.01)
                        for i in range(-3, 0)
                    ) / 3
                    bullish_closes = avg_close_pos >= 0.55
                    at_support     = (
                        (-0.02 <= (price / e20 - 1) <= 0.04) or
                        (-0.01 <= (price / e50 - 1) <= 0.03)
                    )
                    entry_trigger  = bool(
                        len(v) >= 2 and len(c) >= 2 and
                        v[-1] > v[-2] and c[-1] > c[-2]
                    )

                    score = sum([
                        vol_dryup, vol_low5, squeeze3, prior_up,
                        net_accum, weak_selling, rising_lows,
                        bullish_closes, at_support, entry_trigger,
                    ])

                    if score >= 5:
                        high52   = max(h[-252:]) if len(h) >= 252 else max(h)
                        low5     = min(l[-5:]) if len(l) >= 5 else l[-1]
                        buy_lo   = round(price * 0.995, 2)
                        buy_hi   = round(price * 1.005, 2)
                        stop     = round(low5 * 0.99, 2)
                        target   = round(min(price * 1.07, high52 * 0.995), 2)
                        risk     = round((price - stop) / price * 100, 1)
                        reward   = round((target - price) / price * 100, 1)
                        rr       = round(reward / risk, 1) if risk > 0 else 0.0

                        rows.append({
                            "Stock":          sym,
                            "Company":        names.get(sym, sym),
                            "Price (₹)":      round(price, 2),
                            "Mkt Cap (Cr)":   round(mc / 1e7),
                            "Above 200 EMA%": above_200_pct,
                            "Vol Ratio":      vol_ratio,
                            "Prior Uptrend":  "✅" if prior_up      else "❌",
                            "3D Squeeze":     "✅" if squeeze3      else "❌",
                            "Net Accum":      "✅" if net_accum     else "❌",
                            "Rising Lows":    "✅" if rising_lows   else "❌",
                            "Entry Trigger":  "✅" if entry_trigger else "❌",
                            "Score /10":      score,
                            "Buy Zone":       f"₹{buy_lo}–{buy_hi}",
                            "Target (₹)":    target,
                            "Stop (₹)":      stop,
                            "R:R":            rr,
                            "Chart":          f"https://www.tradingview.com/chart/?symbol=NSE:{sym}",
                        })
                except Exception:
                    pass

            st.session_state["up_data"] = rows
            _save_history(rows)
            st.session_state["up_universe_count"] = len(tickers)
            st.session_state["up_ema_count"]      = len(ema200_pass)
            st.session_state["up_mcap_count"]     = len(mcap_pass)

        rows = st.session_state.get("up_data", [])

        if rows:
            df = (
                pd.DataFrame(rows)
                .sort_values(["Score /10", "Vol Ratio", "Mkt Cap (Cr)"],
                             ascending=[False, False, False])
                .head(20)
                .reset_index(drop=True)
            )
            df.index += 1

            u_count  = st.session_state.get("up_universe_count", 0)
            e_count  = st.session_state.get("up_ema_count", 0)
            mc_count = st.session_state.get("up_mcap_count", 0)
            strong   = (df["Score /10"] >= 8).sum()

            c1, c2, c3, c4 = st.columns(4)
            for col, label, val in [
                (c1, "Universe",          f"{u_count:,} stocks"),
                (c2, "Above 200 EMA",     f"{e_count:,} stocks"),
                (c3, "Mkt Cap ≥ 1000 Cr", f"{mc_count:,} stocks"),
                (c4, "High Conviction",   f"{strong} setups (≥ 8/10)"),
            ]:
                col.markdown(f"""<div style="background:white;border-left:3px solid #10b981;border-radius:7px;padding:6px 12px;box-shadow:0 1px 3px rgba(0,0,0,0.07);">
<div style="font-size:0.72rem;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">{label}</div>
<div style="font-size:1.0rem;font-weight:700;color:#0f172a;margin-top:2px;">{val}</div></div>""", unsafe_allow_html=True)

            st.markdown(
                f"**{len(df)} top setups** from {mc_count} qualifying stocks  •  "
                f"**{strong} high-conviction** (Score ≥ 8/10)"
            )

            st.dataframe(
                df,
                use_container_width=True,
                height=min(700, 56 + len(df) * 35),
                column_config={
                    "Stock":          st.column_config.TextColumn("Stock"),
                    "Company":        st.column_config.TextColumn("Company"),
                    "Price (₹)":     st.column_config.NumberColumn("Price (₹)",       format="₹%.2f"),
                    "Mkt Cap (Cr)":  st.column_config.NumberColumn("Mkt Cap (Cr)",    format="₹%,d Cr"),
                    "Above 200 EMA%": st.column_config.NumberColumn("Above 200 EMA",  format="+%.1f%%"),
                    "Vol Ratio":     st.column_config.NumberColumn("Vol Ratio",       format="%.2fx"),
                    "Prior Uptrend": st.column_config.TextColumn("Prior Uptrend"),
                    "3D Squeeze":    st.column_config.TextColumn("3D Squeeze"),
                    "Net Accum":     st.column_config.TextColumn("Net Accum (10D)"),
                    "Rising Lows":   st.column_config.TextColumn("Rising Lows"),
                    "Entry Trigger": st.column_config.TextColumn("Entry Trigger"),
                    "Score /10":     st.column_config.NumberColumn("Score /10",       format="%d ⭐"),
                    "Buy Zone":      st.column_config.TextColumn("Buy Zone"),
                    "Target (₹)":   st.column_config.NumberColumn("Target (₹)",      format="₹%.2f"),
                    "Stop (₹)":     st.column_config.NumberColumn("Stop (₹)",        format="₹%.2f"),
                    "R:R":           st.column_config.NumberColumn("Reward:Risk",     format="%.1f:1"),
                    "Chart":         st.column_config.LinkColumn("TradingView",       display_text="📈 Chart"),
                },
            )

            st.caption(f"📁 Saved to history.xlsx — {len(df)} stocks logged for today ({_dt.date.today()})")

            with st.expander("ℹ️ How Project Up works"):
                st.markdown("""
**Filters applied before scoring:**
- Market Cap > ₹1000 Cr — avoids illiquid small caps
- Price above 200 EMA — in a long-term uptrend

**10 Setup Detection Signals — Score /10:**

| # | Signal | What it checks |
|---|---|---|
| 1 | **Vol Dry-Up** | Today's volume < 70% of 20-day avg — sellers absent |
| 2 | **Vol at 5-Day Low** | Today is the quietest day in 5 sessions — maximum dryness |
| 3 | **3-Day Squeeze** | Last 3 days' range < 4% of price — coiling, spring loaded |
| 4 | **Prior Uptrend** | Close today > close 15 days ago — was rising before this pause |
| 5 | **Strong Net Accum (10D)** | Buying volume > 120% of selling volume over 10 sessions — smart money in |
| 6 | **Weak Selling (5D)** | Down days in last 5 had below-avg volume — no real distribution |
| 7 | **Rising Lows** | Low today > low 4 days ago > low 7 days ago — floor rising |
| 8 | **Avg Bullish Close (3D)** | Buyers consistently closing in top half of range for 3 days |
| 9 | **At Support** | Within ±range of 20 EMA or 50 EMA — sitting on natural support |
| 10 | **Entry Trigger** | Volume expanding AND closing up today — compression ending, move beginning |

**Score /10** — min 5 to appear. High-conviction = 8+.
**Stop ₹** — just below 5-day low (structure-based).
**Target ₹** — lesser of +7% from close or 52-week high.
⚠️ For informational purposes only — not financial advice.
""")
        else:
            st.info("No stocks met all criteria right now. Try again at or after market close (3:30 PM IST).")

# ── TAB 2: BACKTEST ────────────────────────────────────────────────────────────
with tab2:
    st.caption(
        "Checks how past screener picks performed over the next 1, 3, and 5 trading days. "
        "Uses the last 10 pick-dates from history.xlsx."
    )

    if st.button("📊 Run Backtest", key="run_bt"):
        st.session_state.pop("bt_data", None)
        st.session_state["bt_requested"] = True

    if st.session_state.get("bt_requested") and "bt_data" not in st.session_state:
        import yfinance as yf

        hist_path = _os.path.join(_DIR, "history.xlsx")
        if not _os.path.exists(hist_path):
            st.warning("No history.xlsx found. Run the screener first to build pick history.")
            st.session_state.pop("bt_requested", None)
        else:
            raw = pd.read_excel(hist_path, dtype=str)
            raw["Date"]      = pd.to_datetime(raw["Date"], errors="coerce")
            raw["Score /10"] = pd.to_numeric(raw["Score /10"], errors="coerce")
            raw["Price (₹)"] = pd.to_numeric(raw["Price (₹)"], errors="coerce")
            raw = raw.dropna(subset=["Date", "Stock"])

            dates_avail = sorted(raw["Date"].unique())[-10:]
            raw = raw[raw["Date"].isin(dates_avail)]

            bt_tickers = [f"{s}.NS" for s in raw["Stock"].unique().tolist()]
            with st.spinner(f"Downloading forward prices for {len(bt_tickers)} stocks…"):
                try:
                    dh = yf.download(bt_tickers, period="3mo", auto_adjust=True,
                                     progress=False, threads=True)
                    bt_close = {}
                    try:
                        sub = dh["Close"]
                    except KeyError:
                        sub = None
                    if sub is not None:
                        if isinstance(sub, pd.Series):
                            bt_close[bt_tickers[0]] = sub.dropna()
                        else:
                            for col in sub.columns:
                                s = sub[col].dropna()
                                if len(s) > 0:
                                    bt_close[col] = s
                except Exception:
                    bt_close = {}

            today = pd.Timestamp.today().normalize()
            bt_rows = []

            for _, row in raw.iterrows():
                pick_date = row["Date"]
                sym       = row["Stock"]
                tkr       = f"{sym}.NS"

                if tkr not in bt_close:
                    continue

                prices     = bt_close[tkr].sort_index()
                dates_list = list(prices.index)
                # last trading day on or before pick_date (handles weekend picks)
                pick_pos = None
                for i, d in enumerate(dates_list):
                    if d <= pick_date:
                        pick_pos = i
                if pick_pos is None:
                    continue
                pick_px = float(prices.iloc[pick_pos])

                def _fwd(offset):
                    idx = pick_pos + offset
                    if idx >= len(dates_list) or dates_list[idx] > today:
                        return None, None
                    p = float(prices.iloc[idx])
                    return round(p, 2), round((p - pick_px) / pick_px * 100, 2)

                p1, r1 = _fwd(1)
                p3, r3 = _fwd(3)
                p5, r5 = _fwd(5)

                bt_rows.append({
                    "Date":       str(pick_date.date()),
                    "Stock":      sym,
                    "Company":    row.get("Company", sym),
                    "Score":      row.get("Score /10"),
                    "Pick ₹":     round(pick_px, 2),
                    "D+1 ₹":      p1,   "D+1 %":  r1,
                    "D+1":        "✅" if r1 and r1 > 0 else ("❌" if r1 is not None else "—"),
                    "D+3 ₹":      p3,   "D+3 %":  r3,
                    "D+3":        "✅" if r3 and r3 > 0 else ("❌" if r3 is not None else "—"),
                    "D+5 ₹":      p5,   "D+5 %":  r5,
                    "D+5":        "✅" if r5 and r5 > 0 else ("❌" if r5 is not None else "—"),
                })

            st.session_state["bt_data"] = bt_rows

    bt_rows = st.session_state.get("bt_data", [])
    if bt_rows:
        bt_df = pd.DataFrame(bt_rows)
        bt_df["Score"] = pd.to_numeric(bt_df["Score"], errors="coerce")
        pending = bt_df[["D+1 %","D+3 %","D+5 %"]].isnull().all(axis=None)
        if pending:
            st.info("Picks found but all forward prices are still pending — markets haven't traded since the last pick date. Check back after the next trading session.")


        def _hr(series):
            v = series.dropna()
            return f"{(v > 0).mean()*100:.1f}%" if len(v) else "N/A"
        def _avg(series):
            v = series.dropna()
            return f"{v.mean():.2f}%" if len(v) else "N/A"

        st.markdown("#### Summary")
        m1, m2, m3, m4, m5, m6 = st.columns(6)
        for col, label, val in [
            (m1, "Total Picks",   len(bt_df)),
            (m2, "Dates Covered", bt_df["Date"].nunique()),
            (m3, "Hit Rate 1D",   _hr(bt_df["D+1 %"])),
            (m4, "Hit Rate 3D",   _hr(bt_df["D+3 %"])),
            (m5, "Hit Rate 5D",   _hr(bt_df["D+5 %"])),
            (m6, "Avg Return 5D", _avg(bt_df["D+5 %"])),
        ]:
            col.markdown(f"""<div style="background:white;border-left:3px solid #10b981;border-radius:7px;padding:6px 12px;box-shadow:0 1px 3px rgba(0,0,0,0.07);">
<div style="font-size:0.72rem;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;font-weight:600;">{label}</div>
<div style="font-size:1.0rem;font-weight:700;color:#0f172a;margin-top:2px;">{val}</div></div>""", unsafe_allow_html=True)

        hi = bt_df[bt_df["Score"] >= 8]
        lo = bt_df[bt_df["Score"] <  8]
        if len(hi) and len(lo):
            st.markdown("**Score ≥ 8 vs Score 5–7:**")
            sc1, sc2 = st.columns(2)
            sc1.markdown(
                f"**Score ≥ 8** ({len(hi)} picks)  \n"
                f"Hit Rate — 1D: {_hr(hi['D+1 %'])} · 3D: {_hr(hi['D+3 %'])} · 5D: {_hr(hi['D+5 %'])}  \n"
                f"Avg Return — 1D: {_avg(hi['D+1 %'])} · 3D: {_avg(hi['D+3 %'])} · 5D: {_avg(hi['D+5 %'])}"
            )
            sc2.markdown(
                f"**Score 5–7** ({len(lo)} picks)  \n"
                f"Hit Rate — 1D: {_hr(lo['D+1 %'])} · 3D: {_hr(lo['D+3 %'])} · 5D: {_hr(lo['D+5 %'])}  \n"
                f"Avg Return — 1D: {_avg(lo['D+1 %'])} · 3D: {_avg(lo['D+3 %'])} · 5D: {_avg(lo['D+5 %'])}"
            )

        st.markdown("#### Pick-by-Pick Results")
        disp_cols = ["Date","Stock","Company","Score","Pick ₹",
                     "D+1 %","D+1","D+3 %","D+3","D+5 %","D+5"]
        st.dataframe(
            bt_df[[c for c in disp_cols if c in bt_df.columns]],
            use_container_width=True,
            height=min(700, 56 + len(bt_df) * 35),
            column_config={
                "Pick ₹":  st.column_config.NumberColumn("Pick Price", format="₹%.2f"),
                "D+1 ₹":   st.column_config.NumberColumn("D+1 Price",  format="₹%.2f"),
                "D+3 ₹":   st.column_config.NumberColumn("D+3 Price",  format="₹%.2f"),
                "D+5 ₹":   st.column_config.NumberColumn("D+5 Price",  format="₹%.2f"),
                "D+1 %":   st.column_config.NumberColumn("D+1 Return", format="%+.2f%%"),
                "D+3 %":   st.column_config.NumberColumn("D+3 Return", format="%+.2f%%"),
                "D+5 %":   st.column_config.NumberColumn("D+5 Return", format="%+.2f%%"),
                "Score":   st.column_config.NumberColumn("Score /10",  format="%d"),
            },
        )

        bt_path = _os.path.join(_DIR, "backtest_results.xlsx")
        try:
            def _stats(df, label):
                r = {"Group": label, "Picks": len(df)}
                for h, col in [("1D","D+1 %"),("3D","D+3 %"),("5D","D+5 %")]:
                    v = df[col].dropna()
                    r[f"Hit Rate {h}"]   = f"{(v>0).mean()*100:.1f}%" if len(v) else "N/A"
                    r[f"Avg Return {h}"] = f"{v.mean():.2f}%"          if len(v) else "N/A"
                return r

            _RET_COLS = ["D+1 ₹","D+3 ₹","D+5 ₹","D+1 %","D+1","D+3 %","D+3","D+5 %","D+5"]
            _KEY_COLS = ["Date", "Stock"]

            def _upsert_up(existing, new):
                if existing is None or len(existing) == 0:
                    return new.copy()
                for df in [existing, new]:
                    for c in _KEY_COLS:
                        if c in df.columns:
                            df[c] = df[c].astype(str).str.strip()
                key_idx = {(r["Date"], r["Stock"]): i for i, r in existing.iterrows()}
                new_rows = []
                for _, row in new.iterrows():
                    k = (str(row["Date"]), str(row["Stock"]))
                    if k in key_idx:
                        ei = key_idx[k]
                        for col in _RET_COLS:
                            if col in row and col in existing.columns:
                                v = row[col]
                                if v is not None and str(v) not in ("—", "nan", "None", ""):
                                    existing.at[ei, col] = v
                    else:
                        new_rows.append(row.to_dict())
                if new_rows:
                    existing = pd.concat([existing, pd.DataFrame(new_rows)], ignore_index=True)
                return existing

            if _os.path.exists(bt_path):
                try:
                    _existing = pd.read_excel(bt_path, sheet_name="Pick Results", dtype=str)
                    combined_bt = _upsert_up(_existing, bt_df.astype(str))
                except Exception:
                    combined_bt = bt_df.astype(str)
            else:
                combined_bt = bt_df.astype(str)

            for col in ["D+1 %", "D+3 %", "D+5 %"]:
                combined_bt[col] = pd.to_numeric(combined_bt[col], errors="coerce")

            # keep only last 30 days
            combined_bt["Date"] = pd.to_datetime(combined_bt["Date"], errors="coerce")
            cutoff = pd.Timestamp.today() - pd.Timedelta(days=30)
            combined_bt = combined_bt[combined_bt["Date"] >= cutoff]
            combined_bt["Date"] = combined_bt["Date"].dt.strftime("%Y-%m-%d")

            hi_c = combined_bt[pd.to_numeric(combined_bt.get("Score", pd.Series()), errors="coerce") >= 8]
            lo_c = combined_bt[pd.to_numeric(combined_bt.get("Score", pd.Series()), errors="coerce") <  8]

            summary_rows = [_stats(combined_bt, "All Picks")]
            if len(hi_c): summary_rows.append(_stats(hi_c, "Score ≥ 8"))
            if len(lo_c): summary_rows.append(_stats(lo_c, "Score 5–7"))

            _write_formatted_excel(bt_path, {
                "Pick Results": combined_bt,
                "Summary":      pd.DataFrame(summary_rows),
            })
        except Exception:
            pass

        st.caption(f"📁 Saved to backtest_results.xlsx ({len(combined_bt)} total rows)")

st.markdown("</div>", unsafe_allow_html=True)
st.markdown("</div>", unsafe_allow_html=True)
